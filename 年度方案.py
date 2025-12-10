# 第一步：调整导入顺序（Streamlit必须放在最顶部）+ 规范格式
import streamlit as st  # 核心库优先导入
import uuid  # 生成唯一标识（修复注释格式）
import pandas as pd
import numpy as np
import os
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt  # 绘图库

# -------------------------- 全局Session State初始化（统一放在导入后，避免缺失） --------------------------
# 1. 月份选择状态（原有）
if "selected_months" not in st.session_state:
    st.session_state.selected_months = []  # 初始为空，避免状态缺失

# 2. 市场化小时数相关（新增，解决后续手动配置报错）
if "auto_calculate" not in st.session_state:
    st.session_state.auto_calculate = True  # 默认自动计算
if "manual_market_hours_global" not in st.session_state:
    st.session_state.manual_market_hours_global = 0.0  # 全局手动小时数
if "manual_market_hours_monthly" not in st.session_state:
    st.session_state.manual_market_hours_monthly = {month: 0.0 for month in range(1, 13)}  # 分月手动小时数

# 3. 分月参数初始化（含电价、限电率等，避免KeyError）
if "monthly_params" not in st.session_state:
    st.session_state.monthly_params = {
        month: {
            "mechanism_mode": "小时数",
            "mechanism_value": 0.0,
            "guaranteed_mode": "小时数",
            "guaranteed_value": 0.0,
            "power_limit_rate": 0.0,
            "mechanism_price": 0.0,
            "guaranteed_price": 0.0
        } for month in range(1, 13)
    }

# 4. 装机容量/基础数据等核心状态（可选，根据后续代码补充）
if "installed_capacity" not in st.session_state:
    st.session_state.installed_capacity = 0.0  # 装机容量（MW）
if "monthly_data" not in st.session_state:
    st.session_state.monthly_data = {}  # 分月基础数据存储
# 新增：方案电量+基准总量存储（核心，保证数据联动）
if "scheme_power_data" not in st.session_state:
    # 结构：{月份: {"方案一": {"periods": {时段: 电量}, "base_total": 总量}, "方案二": {...}}}
    st.session_state.scheme_power_data = {
        month: {
            "方案一": {"periods": {}, "base_total": 0.0},  # periods=时段电量, base_total=比例调整后的基准总量
            "方案二": {"periods": {}, "base_total": 0.0}
        } for month in range(1, 13)
    }

# -------------------------- 必备：区域-省份映射字典（合并去重，保留详细版本） --------------------------
REGIONS = {
    "总部": ["北京"],
    "华北": ["首都", "河北", "冀北", "山东", "山西", "天津"],
    "华东": ["安徽", "福建", "江苏", "上海", "浙江"],
    "华中": ["湖北", "河南", "湖南", "江西"],
    "东北": ["吉林", "黑龙江", "辽宁", "蒙东"],
    "西北": ["甘肃", "宁夏", "青海", "陕西", "新疆"],
    "西南": ["重庆", "四川", "西藏"],
    "南方": ["广东", "广西", "云南", "海南", "贵州"],
    "内蒙古电网": ["蒙西"]
}

# -------------------------- 全局配置（页面样式） --------------------------
st.set_page_config(
    page_title="新能源电厂年度方案设计系统",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

if "initialized" not in st.session_state:
    # 基础信息默认值
    st.session_state.current_year = 2025
    st.session_state.current_region = "总部"
    st.session_state.current_province = REGIONS["总部"][0]  # 联动区域默认省份
    st.session_state.current_power_plant = "示例电厂"
    st.session_state.current_plant_type = "风电"
    st.session_state.installed_capacity = 0.0
    st.session_state.current_region = "总部"
    st.session_state.current_province = "北京"
    st.session_state.batch_mech_price = 0.0  # 批量-机制电价
    st.session_state.batch_gua_price = 0.0   # 批量-保障性电价
    
    # 光伏套利时段默认配置（首次运行不报错）
    st.session_state["pv_core_start_key"] = 11
    st.session_state["pv_core_end_key"] = 14
    st.session_state["pv_edge_start_key"] = 6
    st.session_state["pv_edge_end_key"] = 18
    
    # 市场化小时数相关（新增分月手动小时数配置）
    st.session_state.auto_calculate = True  # 默认自动计算
    st.session_state.manual_market_hours_global = 0.0  # 全局手动小时数（兼容旧逻辑）
    st.session_state.manual_market_hours_monthly = {month: 0.0 for month in range(1, 13)}  # 分月手动小时数
    
    # 数据存储容器
    st.session_state.monthly_data = {}  # 分月基础数据
    st.session_state.selected_months = []  # 选中的月份
    st.session_state.trade_power_typical = {}  # 方案一结果
    st.session_state.trade_power_arbitrage = {}  # 方案二结果
    st.session_state.market_hours = {}  # 分月市场化小时数
    st.session_state.gen_hours = {}  # 分月发电小时数
    st.session_state.total_annual_trade = 0.0  # 年度总电量
    st.session_state.calculated = False  # 是否已生成方案
    
    # 分月电量参数（每个月独立存储）
    st.session_state.monthly_params = {
        month: {  # 1-12月，每个月对应独立参数
            "mechanism_mode": "小时数",    # 机制电量输入模式
            "mechanism_value": 0.0,        # 机制电量数值
            "guaranteed_mode": "小时数",   # 保障性电量输入模式
            "guaranteed_value": 0.0,       # 保障性电量数值
            "power_limit_rate": 0.0,       # 限电率(%)
            "mechanism_price": 0.0,        # 新增：机制电价(元/MWh)
            "guaranteed_price": 0.0        # 新增：保障性电价(元/MWh)
        } for month in range(1, 13)
    }
    
    # 批量应用的默认参数（用于批量设置时的初始值）
    st.session_state.batch_mech_mode = "小时数"
    st.session_state.batch_mech_value = 0.0
    st.session_state.batch_gua_mode = "小时数"
    st.session_state.batch_gua_value = 0.0
    st.session_state.batch_limit_rate = 0.0
    
    # 标记初始化完成
    st.session_state.initialized = True

# -------------------------- 核心工具函数 --------------------------
def get_days_in_month(year, month):
    """根据年份和月份获取天数（处理闰年）"""
    if month == 2:
        return 29 if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0) else 28
    elif month in [4, 6, 9, 11]:
        return 30
    else:
        return 31

def get_pv_arbitrage_hours():
    """获取光伏套利曲线的时段划分（从session state读取配置）"""
    # 安全获取配置值（转为整数，避免类型错误）
    core_start = int(st.session_state.get("pv_core_start_key", 11))
    core_end = int(st.session_state.get("pv_core_end_key", 14))
    edge_start = int(st.session_state.get("pv_edge_start_key", 6))
    edge_end = int(st.session_state.get("pv_edge_end_key", 18))
    
    # 校验时段有效性（防止超出1-24范围）
    core_start = max(1, min(24, core_start))
    core_end = max(1, min(24, core_end))
    edge_start = max(1, min(24, edge_start))
    edge_end = max(1, min(24, edge_end))
    
    # 确保起始<=结束
    if core_start > core_end:
        core_start, core_end = core_end, core_start
    if edge_start > edge_end:
        edge_start, edge_end = edge_end, edge_start
    
    # 核心时段（中午，电量接收端）
    core_hours = list(range(core_start, core_end + 1))
    # 边缘时段（两端，电量转出端）
    edge_hours = [h for h in range(edge_start, edge_end + 1) if h not in core_hours]
    # 无效时段（非发电时段）
    invalid_hours = [h for h in range(1, 25) if h not in range(edge_start, edge_end + 1)]
    
    return {
        "core": core_hours,       # 中午核心时段
        "edge": edge_hours,       # 两端边缘时段
        "invalid": invalid_hours, # 无效时段
        "config": {
            "core_start": core_start,
            "core_end": core_end,
            "edge_start": edge_start,
            "edge_end": edge_end
        }
    }

def init_month_template(month):
    """初始化单个月份的模板数据"""
    hours = list(range(1, 25))
    return pd.DataFrame({
        "时段": hours,
        "平均发电量(MWh)": [0.0]*24,
        "当月各时段累计发电量(MWh)": [0.0]*24,
        "现货价格(元/MWh)": [0.0]*24,
        "中长期价格(元/MWh)": [0.0]*24,
        "年份": st.session_state.current_year,
        "月份": month,
        "电厂名称": st.session_state.current_power_plant,
        "电厂类型": st.session_state.current_plant_type,
        "区域": st.session_state.current_region,
        "省份": st.session_state.current_province
    })

def export_template():
    """导出Excel模板（包含12个月份子表）"""
    wb = Workbook()
    wb.remove(wb.active)
    for month in range(1, 13):
        ws = wb.create_sheet(title=f"{month}月")
        template_df = init_month_template(month)
        for r in dataframe_to_rows(template_df, index=False, header=True):
            ws.append(r)
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def batch_import_excel(file):
    """批量导入Excel（按子表名称匹配月份）"""
    monthly_data = {}
    try:
        xls = pd.ExcelFile(file)
        for sheet_name in xls.sheet_names:
            if not sheet_name.endswith("月"):
                st.warning(f"跳过无效子表：{sheet_name}（需命名为“1月”-“12月”）")
                continue
            try:
                month = int(sheet_name.replace("月", ""))
                if month < 1 or month > 12:
                    st.warning(f"跳过无效月份子表：{sheet_name}（需1-12月）")
                    continue
                df = pd.read_excel(file, sheet_name=sheet_name)
                required_cols = ["时段", "平均发电量(MWh)", "当月各时段累计发电量(MWh)", "现货价格(元/MWh)", "中长期价格(元/MWh)"]
                if not all(col in df.columns for col in required_cols):
                    st.warning(f"子表{sheet_name}缺少必要列，跳过")
                    continue
                df["年份"] = st.session_state.current_year
                df["电厂名称"] = st.session_state.current_power_plant
                df["电厂类型"] = st.session_state.current_plant_type
                df["区域"] = st.session_state.current_region
                df["省份"] = st.session_state.current_province
                monthly_data[month] = df
            except Exception as e:
                st.warning(f"处理子表{sheet_name}失败：{str(e)}")
        return monthly_data
    except Exception as e:
        st.error(f"批量导入失败：{str(e)}")
        return None

def calculate_core_params_monthly(month, installed_capacity):
    """按月份计算核心参数（内部读取分月参数）"""
    # 安全获取该月份的分月参数（避免KeyError）
    month_params = st.session_state.monthly_params.get(month, {
        "power_limit_rate": 0.0,
        "mechanism_mode": "小时数",
        "mechanism_value": 0.0,
        "guaranteed_mode": "小时数",
        "guaranteed_value": 0.0
    })
    
    # 提取参数（带默认值，防止参数缺失）
    power_limit_rate = month_params.get("power_limit_rate", 0.0)
    mechanism_mode = month_params.get("mechanism_mode", "小时数")
    mechanism_value = month_params.get("mechanism_value", 0.0)
    guaranteed_mode = month_params.get("guaranteed_mode", "小时数")
    guaranteed_value = month_params.get("guaranteed_value", 0.0)
    
    # 校验基础数据是否存在
    if month not in st.session_state.monthly_data:
        st.warning(f"⚠️ 月份{month}无基础数据，发电小时数按0计算")
        return 0.0, 0.0
    
    df = st.session_state.monthly_data[month]
    total_generation = df["当月各时段累计发电量(MWh)"].sum()
    
    # 计算发电小时数（避免装机容量为0）
    gen_hours = round(total_generation / installed_capacity, 2) if installed_capacity > 0 else 0.0
    if gen_hours <= 0:
        st.warning(f"⚠️ 月份{month}发电小时数为0（累计发电量：{total_generation:.2f} MWh，装机容量：{installed_capacity:.2f} MW）")
        market_hours = 0.0
    else:
        # 计算可用小时数（扣除限电）
        available_hours = gen_hours * (1 - power_limit_rate / 100)
        
        # 扣除机制电量
        if mechanism_mode == "小时数":
            available_hours -= mechanism_value
        else:  # 比例(%)
            available_hours -= gen_hours * (mechanism_value / 100)
        
        # 扣除保障性电量
        if guaranteed_mode == "小时数":
            available_hours -= guaranteed_value
        else:  # 比例(%)
            available_hours -= gen_hours * (guaranteed_value / 100)
        
        # 可用小时数不能为负
        available_hours = max(available_hours, 0.0)
        
        # 手动/自动模式区分 + 发电能力校验
        if st.session_state.auto_calculate:
            market_hours = max(round(available_hours, 2), 0.0)
        else:
            # 读取分月手动小时数
            manual_hours = st.session_state.manual_market_hours_monthly.get(month, 0.0)
            # 校验：手动小时数不能超过可用小时数（发电能力上限）
            if manual_hours > available_hours:
                st.warning(f"⚠️ 月份{month}手动小时数({manual_hours})超过可用小时数({available_hours})，已自动截断")
                market_hours = max(round(available_hours, 2), 0.0)
            else:
                market_hours = max(round(manual_hours, 2), 0.0)
    
    return gen_hours, market_hours

def calculate_trade_power_typical(month, market_hours, installed_capacity):
    """方案一：典型出力曲线（按发电权重分配）"""
    # 先校验基础数据
    if month not in st.session_state.monthly_data:
        st.warning(f"⚠️ 月份{month}无基础数据，方案一计算失败")
        return None, 0.0
    
    df = st.session_state.monthly_data[month]
    avg_generation_list = df["平均发电量(MWh)"].tolist()
    total_trade_power = market_hours * installed_capacity
    total_avg_generation = sum(avg_generation_list)
    
    if installed_capacity <= 0 or market_hours <= 0 or total_avg_generation <= 0:
        st.warning(f"⚠️ 月份{month}参数异常（装机容量/市场化小时数/平均发电量不能为0）")
        return None, 0.0
    
    trade_data = []
    for hour, avg_gen in enumerate(avg_generation_list, 1):
        proportion = avg_gen / total_avg_generation
        trade_power = total_trade_power * proportion
        trade_data.append({
            "时段": hour,
            "平均发电量(MWh)": avg_gen,
            "时段比重(%)": round(proportion * 100, 4),
            "方案一月度电量(MWh)": round(trade_power, 2)
        })
    trade_df = pd.DataFrame(trade_data)
    trade_df["年份"] = st.session_state.current_year
    trade_df["月份"] = month
    trade_df["电厂名称"] = st.session_state.current_power_plant
    trade_df = trade_df.fillna(0.0)
    trade_df["方案一月度电量(MWh)"] = trade_df["方案一月度电量(MWh)"].astype(np.float64)
    
    # 校验列是否存在（防止生成失败）
    if "方案一月度电量(MWh)" not in trade_df.columns:
        st.error(f"❌ 月份{month}方案一数据列缺失")
        return None, 0.0
    
    return trade_df, round(total_trade_power, 2)

def calculate_trade_power_arbitrage(month, total_trade_power, typical_df):
    """方案二：光伏套利曲线/风电直线曲线"""
    # 先校验基础数据和典型方案数据
    if month not in st.session_state.monthly_data:
        st.warning(f"⚠️ 月份{month}无基础数据，方案二计算失败")
        return None
    if typical_df is None or typical_df.empty:
        st.warning(f"⚠️ 月份{month}典型方案数据无效，方案二计算失败")
        return None
    
    if st.session_state.current_plant_type == "光伏":
        # 光伏方案二：套利曲线（两端电量转移到中午核心时段）
        pv_hours = get_pv_arbitrage_hours()
        core_hours = pv_hours["core"]
        edge_hours = pv_hours["edge"]
        invalid_hours = pv_hours["invalid"]
        
        # 1. 计算典型曲线中边缘时段的总电量（要转移的电量）
        edge_total = typical_df[typical_df["时段"].isin(edge_hours)]["方案一月度电量(MWh)"].sum()
        # 2. 核心时段数量（避免除以0）
        core_count = len(core_hours) if len(core_hours) > 0 else 1
        # 3. 每个核心时段增加的电量
        core_add = edge_total / core_count
        
        trade_data = []
        for idx, row in typical_df.iterrows():
            hour = row["时段"]
            avg_gen = row["平均发电量(MWh)"]
            
            if hour in invalid_hours:
                # 无效时段：电量=0
                trade_power = 0.0
                proportion = 0.0
            elif hour in edge_hours:
                # 边缘时段：电量=0（全部转移）
                trade_power = 0.0
                proportion = 0.0
            elif hour in core_hours:
                # 核心时段：原典型电量 + 转移电量
                trade_power = row["方案一月度电量(MWh)"] + core_add
                proportion = trade_power / total_trade_power if total_trade_power > 0 else 0.0
            else:
                # 其他时段：保持典型电量
                trade_power = row["方案一月度电量(MWh)"]
                proportion = trade_power / total_trade_power if total_trade_power > 0 else 0.0
            
            trade_data.append({
                "时段": hour,
                "平均发电量(MWh)": avg_gen,
                "时段比重(%)": round(proportion * 100, 4),
                "方案二月度电量(MWh)": round(trade_power, 2)
            })
        
        trade_df = pd.DataFrame(trade_data)
    
    else:
        # 风电方案二：24时段直线平均
        avg_generation_list = st.session_state.monthly_data[month]["平均发电量(MWh)"].tolist()
        hourly_trade = total_trade_power / 24 if total_trade_power > 0 else 0.0
        proportion = 1 / 24
        
        trade_data = []
        for hour, avg_gen in enumerate(avg_generation_list, 1):
            trade_data.append({
                "时段": hour,
                "平均发电量(MWh)": avg_gen,
                "时段比重(%)": round(proportion * 100, 4),
                "方案二月度电量(MWh)": round(hourly_trade, 2)
            })
        trade_df = pd.DataFrame(trade_data)
    
    # 数据清洗和补充
    trade_df["年份"] = st.session_state.current_year
    trade_df["月份"] = month
    trade_df["电厂名称"] = st.session_state.current_power_plant
    trade_df = trade_df.fillna(0.0)
    trade_df["方案二月度电量(MWh)"] = trade_df["方案二月度电量(MWh)"].astype(np.float64)
    
    # 确保方案二总电量和方案一一致（修正浮点数误差）
    if total_trade_power > 0:
        trade_df["方案二月度电量(MWh)"] = trade_df["方案二月度电量(MWh)"] * (total_trade_power / trade_df["方案二月度电量(MWh)"].sum())
    
    # 校验列是否存在
    if "方案二月度电量(MWh)" not in trade_df.columns:
        st.error(f"❌ 月份{month}方案二数据列缺失")
        return None
    
    return trade_df

def decompose_double_scheme(typical_df, arbitrage_df, year, month):
    """双方案日分解（返回四列数据：方案一/二月度+日分解）"""
    days = get_days_in_month(year, month)
    df = pd.DataFrame({
        "时段": typical_df["时段"],
        "方案一月度电量(MWh)": typical_df["方案一月度电量(MWh)"],
        "方案一日分解电量(MWh)": round(typical_df["方案一月度电量(MWh)"] / days, 4) if days > 0 else 0.0,
        "方案二月度电量(MWh)": arbitrage_df["方案二月度电量(MWh)"],
        "方案二日分解电量(MWh)": round(arbitrage_df["方案二月度电量(MWh)"] / days, 4) if days > 0 else 0.0,
        "月份天数": days
    })
    df = df.fillna(0.0)
    return df

def export_annual_plan():
    """导出年度方案Excel（双方案月度+日分解四列数据）"""
    # 第一步：过滤有效月份（仅保留生成方案成功的月份）
    valid_months = []
    for month in st.session_state.selected_months:
        # 校验该月份是否同时存在方案一和方案二的数据，且包含目标列
        if (month in st.session_state.trade_power_typical 
            and month in st.session_state.trade_power_arbitrage
            and not st.session_state.trade_power_typical[month].empty
            and not st.session_state.trade_power_arbitrage[month].empty
            and "方案一月度电量(MWh)" in st.session_state.trade_power_typical[month].columns
            and "方案二月度电量(MWh)" in st.session_state.trade_power_arbitrage[month].columns):
            valid_months.append(month)
        else:
            st.warning(f"⚠️ 跳过无效月份 {month}月（方案数据未生成或列缺失）")
    
    if not valid_months:
        st.error("❌ 无有效方案数据可导出，请先确保所有选中月份的方案生成成功！")
        return None  # 无有效数据时返回None，避免后续报错
    
    wb = Workbook()
    wb.remove(wb.active)
    total_annual = 0.0
    
    # 1. 年度汇总表（双方案总量）
    summary_data = []
    scheme2_note = "套利曲线（两端转中午）" if st.session_state.current_plant_type == "光伏" else "直线曲线（24小时平均）"
    pv_config = get_pv_arbitrage_hours()["config"] if st.session_state.current_plant_type == "光伏" else {}
    
    # 循环有效月份（而非全部选中月份）
    for month in valid_months:
        typical_df = st.session_state.trade_power_typical[month]
        arbitrage_df = st.session_state.trade_power_arbitrage[month]
        
        total_typical = typical_df["方案一月度电量(MWh)"].sum()
        total_arbitrage = arbitrage_df["方案二月度电量(MWh)"].sum()
        total_annual += total_typical
        summary_data.append({
            "年份": st.session_state.current_year,
            "月份": month,
            "电厂名称": st.session_state.current_power_plant,
            "电厂类型": st.session_state.current_plant_type,
            "光伏核心时段": f"{pv_config.get('core_start', '')}-{pv_config.get('core_end', '')}点" if st.session_state.current_plant_type == "光伏" else "-",
            "光伏边缘时段": f"{pv_config.get('edge_start', '')}-{pv_config.get('edge_end', '')}点" if st.session_state.current_plant_type == "光伏" else "-",
            "方案一（典型曲线）总电量(MWh)": total_typical,
            "方案二（{}）总电量(MWh)".format(scheme2_note): total_arbitrage,
            "月份天数": get_days_in_month(st.session_state.current_year, month),
            "市场化小时数": st.session_state.market_hours.get(month, 0.0),
            "占年度比重(%)": round(total_typical / total_annual * 100, 2) if total_annual > 0 else 0.0
        })
    
    # 生成年度汇总表（确保有数据才生成）
    if summary_data:
        summary_df = pd.DataFrame(summary_data)
        ws_summary = wb.create_sheet(title="年度汇总")
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws_summary.append(r)
    else:
        st.error("❌ 无有效汇总数据，导出失败！")
        return None
    
    # 2. 各月份详细表（双方案月度+日分解四列）
    for month in valid_months:
        typical_df = st.session_state.trade_power_typical[month]
        arbitrage_df = st.session_state.trade_power_arbitrage[month]
        
        # 基础数据（安全访问）
        base_df = st.session_state.monthly_data.get(month, None)
        if base_df is None:
            st.warning(f"⚠️ 月份 {month}月 基础数据缺失，跳过详细表")
            continue
        
        # 基础数据列校验
        base_cols = ["时段", "平均发电量(MWh)", "现货价格(元/MWh)", "中长期价格(元/MWh)"]
        if not all(col in base_df.columns for col in base_cols):
            st.warning(f"⚠️ 月份 {month}月 基础数据缺少必要列，跳过详细表")
            continue
        
        # 典型曲线（方案一）- 只保留需要的列
        typical_df_selected = typical_df[["时段", "方案一月度电量(MWh)", "时段比重(%)"]].copy()
        typical_df_selected.rename(columns={"时段比重(%)": "方案一时段比重(%)"}, inplace=True)
        
        # 套利/直线曲线（方案二）- 只保留需要的列
        arbitrage_df_selected = arbitrage_df[["时段", "方案二月度电量(MWh)", "时段比重(%)"]].copy()
        arbitrage_df_selected.rename(columns={"时段比重(%)": "方案二时段比重(%)"}, inplace=True)
        
        # 双方案日分解
        decompose_df = decompose_double_scheme(typical_df, arbitrage_df, st.session_state.current_year, month)
        decompose_df = decompose_df[["时段", "方案一日分解电量(MWh)", "方案二日分解电量(MWh)", "月份天数"]].copy()
        
        # 合并所有数据（按时段关联）
        merged_df = base_df[base_cols].merge(typical_df_selected, on="时段")
        merged_df = merged_df.merge(arbitrage_df_selected, on="时段")
        merged_df = merged_df.merge(decompose_df, on="时段")
        
        # 创建子表
        ws_month = wb.create_sheet(title=f"{month}月详情")
        for r in dataframe_to_rows(merged_df, index=False, header=True):
            ws_month.append(r)
    
    # 3. 方案说明表
    ws_desc = wb.create_sheet(title="方案说明")
    pv_hours = get_pv_arbitrage_hours()
    pv_desc = f"""
    光伏方案二（套利曲线）配置：
    - 核心时段（接收电量）：{pv_hours['core']}点
    - 边缘时段（转出电量）：{pv_hours['edge']}点
    - 无效时段：{pv_hours['invalid']}点
    - 逻辑：将边缘时段的市场化交易电量全部转移至核心时段，总电量保持不变
    """ if st.session_state.current_plant_type == "光伏" else """
    风电方案二（直线曲线）：
    - 逻辑：24小时平均分配市场化交易电量，总电量与典型曲线一致
    """
    desc_content = [
        ["新能源电厂年度交易方案说明"],
        [""],
        ["基础信息："],
        [f"电厂名称：{st.session_state.current_power_plant}"],
        [f"电厂类型：{st.session_state.current_plant_type}"],
        [f"年份：{st.session_state.current_year}"],
        [f"区域：{st.session_state.current_region}"],
        [f"省份：{st.session_state.current_province}"],
        [f"装机容量：{st.session_state.installed_capacity} MW"],
        [""],
        ["方案说明："],
        ["方案一（典型曲线）：按各时段平均发电量权重分配市场化交易电量"],
        [pv_desc],
        [""],
        [f"年度总交易电量（典型方案）：{round(total_annual, 2)} MWh"]
    ]
    for row in desc_content:
        ws_desc.append(row)
    
    # 导出Excel
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# -------------------------- 侧边栏配置 --------------------------
with st.sidebar:
    st.header("⚙️ 基础信息配置")
    
    # 1. 年份选择
    years = list(range(2020, 2031))
    current_year = st.session_state.get("current_year", 2025)
    if current_year not in years:
        current_year = 2025
    st.session_state.current_year = st.selectbox(
        "选择年份", years,
        index=years.index(current_year),
        key="sidebar_year"
    )
    
    # 2. 区域/省份选择（联动+兜底）
    current_region = st.session_state.get("current_region", "总部")
    if current_region not in REGIONS.keys():
        current_region = "总部"
    selected_region = st.selectbox(
        "选择区域",
        list(REGIONS.keys()),
        index=list(REGIONS.keys()).index(current_region),
        key="sidebar_region_select"
    )
    st.session_state.current_region = selected_region
    
    # 省份选择（联动区域）
    provinces = REGIONS[selected_region]
    current_province = st.session_state.get("current_province", provinces[0])
    if current_province not in provinces:
        current_province = provinces[0]
    selected_province = st.selectbox(
        "选择省份",
        provinces,
        index=provinces.index(current_province),
        key="sidebar_province_select"
    )
    st.session_state.current_province = selected_province
    
    # 3. 电厂信息
    st.session_state.current_power_plant = st.text_input(
        "电厂名称",
        value=st.session_state.current_power_plant,
        key="sidebar_power_plant"
    )
    st.session_state.current_plant_type = st.selectbox(
        "电厂类型",
        ["风电", "光伏", "水光互补", "风光互补"],
        index=["风电", "光伏", "水光互补", "风光互补"].index(st.session_state.current_plant_type),
        key="sidebar_plant_type"
    )
    
    # 光伏套利时段配置（仅光伏显示）
    if st.session_state.current_plant_type == "光伏":
        st.subheader("☀️ 光伏套利曲线配置")
        st.write("核心时段（中午，接收电量）")
        col_pv1, col_pv2 = st.columns(2)
        with col_pv1:
            st.number_input(
                "核心起始（点）", min_value=1, max_value=24,
                value=st.session_state["pv_core_start_key"],
                key="input_pv_core_start"
            )
        with col_pv2:
            st.number_input(
                "核心结束（点）", min_value=1, max_value=24,
                value=st.session_state["pv_core_end_key"],
                key="input_pv_core_end"
            )
        
        st.write("边缘时段（两端，转出电量）")
        col_pv3, col_pv4 = st.columns(2)
        with col_pv3:
            st.number_input(
                "边缘起始（点）", min_value=1, max_value=24,
                value=st.session_state["pv_edge_start_key"],
                key="input_pv_edge_start"
            )
        with col_pv4:
            st.number_input(
                "边缘结束（点）", min_value=1, max_value=24,
                value=st.session_state["pv_edge_end_key"],
                key="input_pv_edge_end"
            )
        
        # 同步input值到session state
        st.session_state["pv_core_start_key"] = st.session_state.get("input_pv_core_start", 11)
        st.session_state["pv_core_end_key"] = st.session_state.get("input_pv_core_end", 14)
        st.session_state["pv_edge_start_key"] = st.session_state.get("input_pv_edge_start", 6)
        st.session_state["pv_edge_end_key"] = st.session_state.get("input_pv_edge_end", 18)
        
        # 显示时段划分
        pv_hours = get_pv_arbitrage_hours()
        st.info(f"""
        时段划分：
        - 核心时段（接收）：{pv_hours['core']}点
        - 边缘时段（转出）：{pv_hours['edge']}点
        - 无效时段：{pv_hours['invalid']}点
        """)
    
    # 4. 装机容量
    installed_capacity = st.number_input(
        "装机容量(MW)", min_value=0.0, value=st.session_state.installed_capacity, step=0.1,
        key="sidebar_installed_capacity", help="电厂总装机容量，单位：兆瓦"
    )
    st.session_state.installed_capacity = installed_capacity  # 同步到session state
    
    # 5. 市场化交易小时数（简化版）
    st.write("#### 市场化交易小时数")
    auto_calculate = st.toggle(
        "自动计算", value=st.session_state.auto_calculate,
        key="sidebar_auto_calculate"
    )
    st.session_state.auto_calculate = auto_calculate

    if not st.session_state.auto_calculate:
        # 仅保留全局手动值（用于批量应用）
        st.session_state.manual_market_hours_global = st.number_input(
            "全局手动值（可批量应用到所有月份）", min_value=0.0, max_value=1000000.0,
            value=st.session_state.manual_market_hours_global, step=0.1,
            key="sidebar_market_hours_global"
        )
        # 批量应用全局值按钮
        if st.button("📌 全局值批量应用到所有月份", key="batch_manual_hours"):
            for month in range(1, 13):
                st.session_state.manual_market_hours_monthly[month] = st.session_state.manual_market_hours_global
            st.success("✅ 已将全局值同步到所有月份！")

# -------------------------- 主页面：电量参数配置 --------------------------
st.subheader("⚡ 电量参数配置")

# 1. 批量应用参数（一键同步到所有月份）
st.write("#### 批量应用（同步到所有月份）")
col_mech1, col_mech2 = st.columns([2, 1])
with col_mech1:
    st.session_state.batch_mech_mode = st.selectbox(
        "机制电量输入模式", ["小时数", "比例(%)"],
        index=0 if st.session_state.batch_mech_mode == "小时数" else 1,
        key="batch_mech_mode_sel"
    )
with col_mech2:
    mech_max = 100.0 if st.session_state.batch_mech_mode == "比例(%)" else 1000000.0
    st.session_state.batch_mech_value = st.number_input(
        "机制电量数值", min_value=0.0, max_value=mech_max, 
        value=st.session_state.batch_mech_value, step=0.1,
        key="batch_mech_val_inp"
    )

col_gua1, col_gua2 = st.columns([2, 1])
with col_gua1:
    st.session_state.batch_gua_mode = st.selectbox(
        "保障性电量输入模式", ["小时数", "比例(%)"],
        index=0 if st.session_state.batch_gua_mode == "小时数" else 1,
        key="batch_gua_mode_sel"
    )
with col_gua2:
    gua_max = 100.0 if st.session_state.batch_gua_mode == "比例(%)" else 1000000.0
    st.session_state.batch_gua_value = st.number_input(
        "保障性电量数值", min_value=0.0, max_value=gua_max,
        value=st.session_state.batch_gua_value, step=0.1,
        key="batch_gua_val_inp"
    )

# -------------------------- 新增：机制电价输入（直接插入这里）--------------------------
col_mech_price1, col_mech_price2 = st.columns([2, 1])
with col_mech_price1:
    st.write("机制电价（元/MWh）")
with col_mech_price2:
    st.session_state.batch_mech_price = st.number_input(
        "机制电价数值", min_value=0.0,
        value=st.session_state.batch_mech_price, step=0.1,
        key="batch_mech_price_inp"
    )

# -------------------------- 新增：保障性电价输入（直接插入这里）--------------------------
col_gua_price1, col_gua_price2 = st.columns([2, 1])
with col_gua_price1:
    st.write("保障性电价（元/MWh）")
with col_gua_price2:
    st.session_state.batch_gua_price = st.number_input(
        "保障性电价数值", min_value=0.0,
        value=st.session_state.batch_gua_price, step=0.1,
        key="batch_gua_price_inp"
    )

# 原有的限电率输入（保持不变）
st.session_state.batch_limit_rate = st.number_input(
    "限电率(%)", min_value=0.0, max_value=100.0,
    value=st.session_state.batch_limit_rate, step=0.1,
    key="batch_limit_rate_inp"
)

# 批量应用按钮（同步新增的电价参数，修改这里的字典）
if st.button("📌 一键应用到所有月份", type="primary", key="batch_apply_btn"):
    for month in range(1, 13):
        st.session_state.monthly_params[month] = {
            "mechanism_mode": st.session_state.batch_mech_mode,
            "mechanism_value": st.session_state.batch_mech_value,
            "guaranteed_mode": st.session_state.batch_gua_mode,
            "guaranteed_value": st.session_state.batch_gua_value,
            "power_limit_rate": st.session_state.batch_limit_rate,
            "mechanism_price": st.session_state.batch_mech_price,  # 新增：同步机制电价
            "guaranteed_price": st.session_state.batch_gua_price    # 新增：同步保障性电价
        }
    st.success("✅ 已将当前参数（含电价）同步到所有月份！")

# 先在分月调整Expander内部、selected_month定义后生成唯一前缀（核心）
with st.expander("🔧 分月参数调整（单独修改）", expanded=False):
    # 生成8位唯一前缀，彻底避免Key重复
    unique_prefix = str(uuid.uuid4())[:8]
    selected_month = st.selectbox("选择要修改的月份", range(1, 13), key=f"{unique_prefix}_month_param_sel")
    current_params = st.session_state.monthly_params[selected_month]

    # 分月-机制电量（修复Key+删除重复定义）
    st.write(f"##### {selected_month}月 · 机制电量")
    col_m1, col_m2 = st.columns([2, 1])
    with col_m1:
        mech_mode = st.selectbox(
            "输入模式", ["小时数", "比例(%)"],
            index=0 if current_params["mechanism_mode"] == "小时数" else 1,
            key=f"{unique_prefix}_mech_mode_{selected_month}"  # 唯一Key
        )
    with col_m2:
        m_max = 100.0 if mech_mode == "比例(%)" else 1000000.0
        mech_val = st.number_input(
            "数值", min_value=0.0, max_value=m_max,
            value=current_params["mechanism_value"], step=0.1,
            key=f"{unique_prefix}_mech_val_{selected_month}"  # 唯一Key
        )

    # 分月-保障性电量（修复Key+删除重复定义）
    st.write(f"##### {selected_month}月 · 保障性电量")
    col_g1, col_g2 = st.columns([2, 1])
    with col_g1:
        gua_mode = st.selectbox(
            "输入模式", ["小时数", "比例(%)"],
            index=0 if current_params["guaranteed_mode"] == "小时数" else 1,
            key=f"{unique_prefix}_gua_mode_{selected_month}"  # 唯一Key
        )
    with col_g2:
        g_max = 100.0 if gua_mode == "比例(%)" else 1000000.0
        gua_val = st.number_input(
            "数值", min_value=0.0, max_value=g_max,
            value=current_params["guaranteed_value"], step=0.1,
            key=f"{unique_prefix}_gua_val_{selected_month}"  # 唯一Key
        )

    # 分月-机制电价（修复Key+删除重复定义）
    st.write(f"##### {selected_month}月 · 机制电价")
    mech_price = st.number_input(
        "机制电价（元/MWh）", min_value=0.0,
        value=current_params["mechanism_price"], step=0.1,
        key=f"{unique_prefix}_mech_price_{selected_month}"  # 唯一Key
    )

    # 分月-保障性电价（修复Key+删除重复定义）
    st.write(f"##### {selected_month}月 · 保障性电价")
    gua_price = st.number_input(
        "保障性电价（元/MWh）", min_value=0.0,
        value=current_params["guaranteed_price"], step=0.1,
        key=f"{unique_prefix}_gua_price_{selected_month}"  # 唯一Key
    )

    # 分月-限电率（补充完整，你原代码漏了这部分定义）
    st.write(f"##### {selected_month}月 · 限电率")
    limit_rate = st.number_input(
        "限电率(%)", min_value=0.0, max_value=100.0,
        value=current_params["power_limit_rate"], step=0.1,
        key=f"{unique_prefix}_limit_rate_{selected_month}"  # 唯一Key
    )

    # 新增：分月-手动市场化小时数（仅手动模式显示，修复语法错误）
    if not st.session_state.auto_calculate:
        st.write(f"##### {selected_month}月 · 手动市场化小时数")
        current_manual_hours = st.session_state.manual_market_hours_monthly.get(selected_month, 0.0)
        manual_hours = st.number_input(
            "市场化小时数（自动校验不超过可用小时数）", 
            min_value=0.0,
            value=current_manual_hours, 
            step=0.1,
            key=f"{unique_prefix}_manual_market_hours_{selected_month}",  # 唯一Key+修复语法
            help="手动设置的小时数不能超过扣除限电/机制/保障性电量后的可用小时数"
        )
        st.session_state.manual_market_hours_monthly[selected_month] = manual_hours

    # 保存按钮（唯一Key）
    col_save, col_empty = st.columns([1, 5])
    with col_save:
        if st.button(
            f"💾 保存{selected_month}月参数", 
            key=f"{unique_prefix}_save_{selected_month}_param",  # 唯一Key
            type="primary"
        ):
            # 保存逻辑（补充manual_hours不影响保存，仅市场化小时数用）
            st.session_state.monthly_params[selected_month] = {
                "mechanism_mode": mech_mode,
                "mechanism_value": mech_val,
                "guaranteed_mode": gua_mode,
                "guaranteed_value": gua_val,
                "power_limit_rate": limit_rate,
                "mechanism_price": mech_price,
                "guaranteed_price": gua_price
            }
            # 可选：提示可用小时数
            if st.session_state.installed_capacity > 0 and selected_month in st.session_state.monthly_data:
                temp_gen_hours, temp_available_hours = calculate_core_params_monthly(selected_month, st.session_state.installed_capacity)
                st.info(f"💡 该月份可用小时数上限：{temp_available_hours:.2f}")
            st.success(f"✅ 已保存{selected_month}月的参数（含电价）！")
            st.rerun()

    # 所有月份参数预览表格（修复格式+新增列）
    st.divider()
    st.write("#### 所有月份参数预览（含电价）")
    param_preview = []  # 初始化列表（避免重复追加）
    for month in range(1, 13):
        p = st.session_state.monthly_params[month]
        param_preview.append({
            "月份": f"{month}月",
            "机制电量": f"{p['mechanism_mode']} · {p['mechanism_value']:.2f}",
            "保障性电量": f"{p['guaranteed_mode']} · {p['guaranteed_value']:.2f}",
            "机制电价(元/MWh)": f"{p['mechanism_price']:.2f}",
            "保障性电价(元/MWh)": f"{p['guaranteed_price']:.2f}",
            "限电率": f"{p['power_limit_rate']:.2f}%",
            "手动市场化小时数": f"{st.session_state.manual_market_hours_monthly.get(month, 0.0):.2f}"
        })
    preview_df = pd.DataFrame(param_preview)
    st.dataframe(preview_df, use_container_width=True, hide_index=True)

# -------------------------- 主页面内容 --------------------------
st.title("⚡ 新能源电厂年度方案设计系统")
scheme2_title = "套利曲线（光伏）/直线曲线（风电）"
st.subheader(
    f"当前配置：{st.session_state.current_year}年 | {st.session_state.current_region} | {st.session_state.current_province} | "
    f"{st.session_state.current_plant_type} | {st.session_state.current_power_plant}"
)
st.caption(f"方案一：典型出力曲线 | 方案二：{scheme2_title}")

# -------------------------- 模板导出与批量导入（合并重复模块，只保留1次）--------------------------
st.divider()
st.header("📤 模板导出与批量导入")
col_import1, col_import2, col_import3 = st.columns(3)

# 1. 导出模板按钮（只保留1个）
with col_import1:
    template_output = export_template()
    st.download_button(
        "📥 导出Excel模板（含12个月）",
        data=template_output,
        file_name=f"{st.session_state.current_power_plant}_{st.session_state.current_year}年方案模板.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

# 2. 批量导入按钮（只保留1个）
with col_import2:
    batch_file = st.file_uploader(
        "📥 批量导入Excel（含多月份子表）",
        type=["xlsx"],
        key="batch_import_file_unique"  # 加unique确保key不重复
    )
    if batch_file is not None:
        monthly_data = batch_import_excel(batch_file)
        if monthly_data:
            st.session_state.monthly_data = monthly_data
            st.session_state.selected_months = sorted(list(monthly_data.keys()))
            st.success(f"✅ 批量导入成功！共导入{len(monthly_data)}个月份数据")

# 3. 月份选择（终极修复：强制同步，全选后下拉框自动填充12月）
with col_import3:
    st.subheader("选择需要处理的月份", divider="gray")
    
    # 全选/取消全选按钮（点击后直接更新状态+强制刷新）
    col_btn1, col_btn2 = st.columns([1, 1], gap="small")
    with col_btn1:
        if st.button("📅 全选1-12月", key="select_all_final", type="primary", use_container_width=True):
            st.session_state.selected_months = list(range(1, 13))
            st.rerun()  # 强制刷新页面，让下拉框重新渲染
    with col_btn2:
        if st.button("❌ 取消全选", key="deselect_all_final", use_container_width=True):
            st.session_state.selected_months = []
            st.rerun()  # 强制刷新页面
    
    # 手动微调区域（强制绑定session_state，无延迟）
    st.write("### 手动微调（可取消个别月份）")
    # 直接将multiselect的选项和默认值设为session_state的最新值
    manual_selected = st.multiselect(
        label=f"当前已选：{len(st.session_state.selected_months)}个月份",
        options=list(range(1, 13)),  # 固定选项：1-12月
        default=st.session_state.selected_months,  # 强制取当前选中的月份
        key=f"month_multiselect_{len(st.session_state.selected_months)}",  # 用选中数量做Key，强制重新渲染
        format_func=lambda x: f"{x}月",
        placeholder="全选后自动填充12个月"  # 仅在未选时显示提示
    )
    
    # 同步状态（手动调整后更新session_state）
    if manual_selected != st.session_state.selected_months:
        st.session_state.selected_months = manual_selected
        st.rerun()  # 调整后也刷新，确保显示一致
    
    # 状态提示（严格反映最终选中）
    if st.session_state.selected_months:
        months_text = "、".join([f"{m}月" for m in sorted(st.session_state.selected_months)])
        st.info(f"📌 最终选中：{months_text}（共{len(st.session_state.selected_months)}个月份）")
    else:
        st.warning("⚠️ 请选择需要处理的月份（可点击「全选1-12月」快速选择）")

# 二、数据操作按钮
st.divider()
st.header("🔧 数据操作")
col_data1, col_data2, col_data3 = st.columns(3)

# 1. 初始化选中月份模板
with col_data1:
    if st.button("📋 初始化选中月份模板", use_container_width=True, key="init_selected_months"):
        if not st.session_state.selected_months:
            st.warning("⚠️ 请先选择月份")
        else:
            for month in st.session_state.selected_months:
                st.session_state.monthly_data[month] = init_month_template(month)
            st.success(f"✅ 已初始化{len(st.session_state.selected_months)}个月份模板")

# 2. 生成年度双方案（重点修复：严格过滤无效数据）
with col_data2:
    if st.button("📝 生成年度双方案", use_container_width=True, type="primary", key="generate_annual_plan"):
        if not st.session_state.selected_months or not st.session_state.monthly_data:
            st.warning("⚠️ 请先导入/初始化月份数据并选择月份")
        elif st.session_state.installed_capacity <= 0:
            st.warning("⚠️ 请填写有效的装机容量（>0）")
        else:
            with st.spinner("🔄 正在计算年度双方案..."):
                try:
                    trade_typical = {}
                    trade_arbitrage = {}
                    market_hours = {}
                    gen_hours = {}
                    total_annual = 0.0
                    valid_calculated_months = []  # 记录成功计算的月份
                    
                    for month in st.session_state.selected_months:
                        # 计算核心参数（仅传2个参数，内部读取分月参数）
                        if st.session_state.auto_calculate:
                            gh, mh = calculate_core_params_monthly(month, st.session_state.installed_capacity)
                        else:
                            # 手动模式：发电小时数按分月参数计算，市场化小时数用手动输入
                            gh, _ = calculate_core_params_monthly(month, st.session_state.installed_capacity)
                            mh = st.session_state.manual_market_hours
                        
                        # 校验市场化小时数有效性
                        if mh <= 0:
                            st.warning(f"⚠️ 月份{month}市场化小时数为0，跳过该月份")
                            continue
                        
                        market_hours[month] = mh   
                        gen_hours[month] = gh
                        
                        # 方案一：典型曲线（校验返回结果）
                        typical_df, total_typical = calculate_trade_power_typical(month, mh, st.session_state.installed_capacity)
                        if typical_df is None or typical_df.empty or "方案一月度电量(MWh)" not in typical_df.columns:
                            st.error(f"❌ 月份{month}典型方案计算失败，跳过该月份")
                            continue
                        
                        # 方案二：光伏套利/风电直线（校验返回结果）
                        arbitrage_df = calculate_trade_power_arbitrage(month, total_typical, typical_df)
                        if arbitrage_df is None or arbitrage_df.empty or "方案二月度电量(MWh)" not in arbitrage_df.columns:
                            st.error(f"❌ 月份{month}方案二计算失败，跳过该月份")
                            continue
                        
                        # 只有两个方案都成功才存入会话状态
                        trade_typical[month] = typical_df
                        trade_arbitrage[month] = arbitrage_df
                        total_annual += total_typical
                        valid_calculated_months.append(month)
                    
                    # 只有有有效计算结果才更新会话状态
                    if valid_calculated_months:
                        st.session_state.trade_power_typical = trade_typical
                        st.session_state.trade_power_arbitrage = trade_arbitrage
                        st.session_state.market_hours = market_hours
                        st.session_state.gen_hours = gen_hours
                        st.session_state.total_annual_trade = total_annual
                        st.session_state.calculated = True
                        
                        st.success(
                            f"✅ 年度双方案生成成功！\n"
                            f"成功计算月份：{', '.join([f'{m}月' for m in valid_calculated_months])}\n"
                            f"年度总交易电量：{round(total_annual, 2)} MWh"
                        )
                    else:
                        st.error("❌ 所有选中月份的方案计算均失败，请检查基础数据和参数配置！")
                        st.session_state.calculated = False  # 标记为未计算成功
                    
                except Exception as e:
                    st.error(f"❌ 生成方案失败：{str(e)}")
                    st.session_state.calculated = False

# 3. 导出年度方案
with col_data3:
    if st.session_state.calculated and st.session_state.trade_power_typical:
        annual_output = export_annual_plan()
        if annual_output:  # 确保有有效数据才显示下载按钮
            st.download_button(
                "💾 导出年度方案（双方案+日分解）",
                data=annual_output,
                file_name=f"{st.session_state.current_power_plant}_{st.session_state.current_year}年双方案交易数据.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.button(
            "💾 导出年度方案（双方案+日分解）",
            use_container_width=True,
            disabled=True,
            help="请先生成有效的年度方案"
        )

# 三、选中月份数据编辑
if st.session_state.selected_months and st.session_state.monthly_data:
    st.divider()
    st.header("📊 选中月份数据编辑")
    edit_month = st.selectbox(
        "选择要编辑的月份",
        st.session_state.selected_months,
        key="edit_month_select"
    )
    if edit_month in st.session_state.monthly_data:
        edit_df = st.data_editor(
            st.session_state.monthly_data[edit_month],
            column_config={
                "时段": st.column_config.NumberColumn("时段", disabled=True),
                "平均发电量(MWh)": st.column_config.NumberColumn("平均发电量(MWh)", min_value=0.0, step=0.1),
                "当月各时段累计发电量(MWh)": st.column_config.NumberColumn("当月各时段累计发电量(MWh)", min_value=0.0, step=0.1),
                "现货价格(元/MWh)": st.column_config.NumberColumn("现货价格(元/MWh)", min_value=0.0, step=0.1),
                "中长期价格(元/MWh)": st.column_config.NumberColumn("中长期价格(元/MWh)", min_value=0.0, step=0.1),
                "年份": st.column_config.NumberColumn("年份", disabled=True),
                "月份": st.column_config.NumberColumn("月份", disabled=True),
                "电厂名称": st.column_config.TextColumn("电厂名称", disabled=True),
                "电厂类型": st.column_config.TextColumn("电厂类型", disabled=True),
                "区域": st.column_config.TextColumn("区域", disabled=True),
                "省份": st.column_config.TextColumn("省份", disabled=True)
            },
            use_container_width=True,
            num_rows="fixed",
            key=f"edit_df_{edit_month}"
        )
        st.session_state.monthly_data[edit_month] = edit_df

# 四、年度方案展示（重点修复：只展示有效月份）
if st.session_state.calculated and st.session_state.trade_power_typical:
    st.divider()
    st.header(f"📈 {st.session_state.current_year}年度方案展示（双方案对比）")
    
    # 过滤有效展示月份（存在于两个方案中且列名齐全）
    valid_display_months = [
        month for month in st.session_state.selected_months
        if month in st.session_state.trade_power_typical
        and month in st.session_state.trade_power_arbitrage
        and not st.session_state.trade_power_typical[month].empty
        and not st.session_state.trade_power_arbitrage[month].empty
        and "方案一月度电量(MWh)" in st.session_state.trade_power_typical[month].columns
        and "方案二月度电量(MWh)" in st.session_state.trade_power_arbitrage[month].columns
    ]
    
    if not valid_display_months:
        st.warning("⚠️ 无有效方案数据可展示，请重新生成方案")
    else:
        # 1. 年度汇总
        st.subheader("1. 年度汇总")
        summary_data = []
        scheme2_note = "套利曲线" if st.session_state.current_plant_type == "光伏" else "直线曲线"
        for month in valid_display_months:
            typical_df = st.session_state.trade_power_typical[month]
            arbitrage_df = st.session_state.trade_power_arbitrage[month]
            
            typical_total = typical_df["方案一月度电量(MWh)"].sum()
            arbitrage_total = arbitrage_df["方案二月度电量(MWh)"].sum()
            days = get_days_in_month(st.session_state.current_year, month)
            summary_data.append({
                "月份": f"{month}月",
                "月份天数": days,
                "市场化小时数": st.session_state.market_hours.get(month, 0.0),
                "预估发电小时数": st.session_state.gen_hours.get(month, 0.0),
                "方案一总电量(MWh)": typical_total,
                "方案二总电量(MWh)": arbitrage_total,
                "方案二类型": scheme2_note,
                "占年度比重(%)": round(typical_total / st.session_state.total_annual_trade * 100, 2)
            })
        summary_df = pd.DataFrame(summary_data)
        st.dataframe(summary_df, use_container_width=True, hide_index=True)
        st.metric("年度总交易电量（方案一）", f"{st.session_state.total_annual_trade:.2f} MWh")
        
        # 2. 月份方案详情
        st.subheader("2. 月份方案详情（双方案对比）")
        view_month = st.selectbox(
            "选择查看的月份",
            valid_display_months,
            key="view_month_select"
        )
        
        try:
            # 保持联动逻辑：实时读最新数据
            typical_df = st.session_state.trade_power_typical.get(view_month, pd.DataFrame())
            required_cols = ["时段", "方案一月度电量(MWh)"]
            if typical_df.empty or not all(col in typical_df.columns for col in required_cols):
                st.info("⚠️ 暂无有效方案一数据（数据为空或缺少必要列）")
                pass
            else:
                base_df = st.session_state.monthly_data.get(view_month, None)
                if base_df is None or base_df.empty:
                    st.info("⚠️ 缺少基础价格数据，仅展示交易量图表")
                    # 纯交易量交互式柱状图
                    import plotly.express as px
                    fig = px.bar(
                        typical_df,
                        x="时段",
                        y="方案一月度电量(MWh)",
                        title=f"{view_month}月 方案一交易量",
                        labels={"方案一月度电量(MWh)": "交易量（MWh）", "时段": "时段（点）"},
                        color_discrete_sequence=["#4299e1"],  # 柔和蓝色
                        height=350
                    )
                    # 视觉优化：去除背景网格、调整字体
                    fig.update_layout(
                        plot_bgcolor="white",
                        xaxis_showgrid=False,
                        yaxis_showgrid=True,
                        yaxis_gridcolor="#f0f0f0",
                        font=dict(family="Arial", size=11),
                        title_font=dict(size=13, weight="bold"),
                        margin=dict(l=10, r=10, t=30, b=10)  # 紧凑边距
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    # 准备数据（确保长度一致）
                    merged_data = typical_df[["时段", "方案一月度电量(MWh)"]].copy()
                    if len(base_df) >= 24:
                        merged_data["现货价格"] = base_df["现货价格(元/MWh)"].head(24).values
                        merged_data["中长期价格"] = base_df["中长期价格(元/MWh)"].head(24).values
                    else:
                        merged_data["现货价格"] = 0.0
                        merged_data["中长期价格"] = 0.0

                    # 用 Plotly 创建双轴交互式图表
                    import plotly.graph_objects as go
                    fig = go.Figure()

                    # 1. 交易量柱状图（左轴）
                    fig.add_trace(go.Bar(
                        x=merged_data["时段"],
                        y=merged_data["方案一月度电量(MWh)"],
                        name="方案一交易量",
                        yaxis="y1",
                        marker_color="#4299e1",  # 柔和蓝
                        opacity=0.8,
                        hovertemplate="时段：%{x}点<br>交易量：%{y:.2f} MWh<extra></extra>"
                    ))

                    # 2. 现货价格折线（右轴）
                    fig.add_trace(go.Scatter(
                        x=merged_data["时段"],
                        y=merged_data["现货价格"],
                        name="现货价格",
                        yaxis="y2",
                        mode="lines+markers",
                        line=dict(color="#9f7aea", width=2),  # 柔和紫
                        marker=dict(size=4),
                        hovertemplate="时段：%{x}点<br>现货价格：%{y:.2f} 元/MWh<extra></extra>"
                    ))

                    # 3. 中长期价格折线（右轴）
                    fig.add_trace(go.Scatter(
                        x=merged_data["时段"],
                        y=merged_data["中长期价格"],
                        name="中长期价格",
                        yaxis="y2",
                        mode="lines+markers",
                        line=dict(color="#38b2ac", width=2),  # 柔和青
                        marker=dict(size=4),
                        hovertemplate="时段：%{x}点<br>中长期价格：%{y:.2f} 元/MWh<extra></extra>"
                    ))

                    # 视觉+布局优化（核心！）
                    fig.update_layout(
                        # 标题
                        title=f"{view_month}月 方案一交易量与价格对比",
                        title_font=dict(size=13, weight="bold", family="Arial"),
                        title_x=0.5,  # 居中
                        # 背景
                        plot_bgcolor="white",
                        paper_bgcolor="white",
                        # 双轴设置
                        yaxis1=dict(
                            title="交易量（MWh）",
                            title_font=dict(color="#4299e1"),
                            tickfont=dict(color="#4299e1"),
                            gridcolor="#f0f0f0"  # 淡灰网格
                        ),
                        yaxis2=dict(
                            title="价格（元/MWh）",
                            title_font=dict(color="#9f7aea"),
                            tickfont=dict(color="#9f7aea"),
                            overlaying="y",
                            side="right",
                            gridcolor="rgba(0,0,0,0)"  # 隐藏右轴网格，避免重叠
                        ),
                        # 图例
                        legend=dict(
                            orientation="h",  # 水平排列
                            yanchor="bottom",
                            y=-0.2,  # 放在图表下方，不挡数据
                            xanchor="center",
                            x=0.5
                        ),
                        # 边距（紧凑不浪费空间）
                        margin=dict(l=20, r=20, t=30, b=60),
                        # x轴优化
                        xaxis=dict(
                            title="时段（点）",
                            tickmode="array",
                            tickvals=merged_data["时段"],  # 显示所有24时段
                            gridcolor="#f0f0f0"
                        )
                    )

                    # 在 Streamlit 中显示（支持交互）
                    st.plotly_chart(fig, use_container_width=True)

        except Exception as e:
            st.warning(f"📊 方案一图表生成失败：{str(e)}（不影响数据导出）")
            
        # 方案二展示
        st.write(f"### 方案二：{scheme2_note}（{view_month}月）")
        arbitrage_df = st.session_state.trade_power_arbitrage[view_month][["时段", "平均发电量(MWh)", "时段比重(%)", "方案二月度电量(MWh)"]].copy()
        arbitrage_df = arbitrage_df.fillna(0.0)
        arbitrage_df["方案二月度电量(MWh)"] = arbitrage_df["方案二月度电量(MWh)"].astype(np.float64)
        arbitrage_df = arbitrage_df.reset_index(drop=True)
        st.dataframe(arbitrage_df, use_container_width=True, hide_index=True)
        
        # 方案二说明
        if st.session_state.current_plant_type == "光伏":
            pv_hours = get_pv_arbitrage_hours()
            edge_total = typical_df[typical_df["时段"].isin(pv_hours["edge"])]["方案一月度电量(MWh)"].sum()
            core_avg_add = edge_total / len(pv_hours["core"]) if len(pv_hours["core"]) > 0 else 0
            st.info(f"""
            光伏套利曲线说明：
            - 转出时段：{pv_hours['edge']}点（总转出电量={edge_total:.2f} MWh）
            - 接收时段：{pv_hours['core']}点（每时段增加={core_avg_add:.2f} MWh）
            - 总电量：{arbitrage_df['方案二月度电量(MWh)'].sum():.2f} MWh（与方案一一致）
            """)
        else:
            st.info(f"""
            风电直线曲线说明：
            - 24时段平均分配，每时段电量={arbitrage_df['方案二月度电量(MWh)'].iloc[0]:.2f} MWh
            - 总电量：{arbitrage_df['方案二月度电量(MWh)'].sum():.2f} MWh（与方案一一致）
            """)
        
        try:
            # 保持联动逻辑：实时读最新数据
            arbitrage_df = st.session_state.trade_power_arbitrage.get(view_month, pd.DataFrame())
            required_cols = ["时段", "方案二月度电量(MWh)"]
            if arbitrage_df.empty or not all(col in arbitrage_df.columns for col in required_cols):
                st.info("⚠️ 暂无有效方案二数据（数据为空或缺少必要列）")
                pass
            else:
                base_df = st.session_state.monthly_data.get(view_month, None)
                if base_df is None or base_df.empty:
                    st.info("⚠️ 缺少基础价格数据，仅展示交易量图表")
                    import plotly.express as px
                    fig = px.bar(
                        arbitrage_df,
                        x="时段",
                        y="方案二月度电量(MWh)",
                        title=f"{view_month}月 方案二交易量",
                        labels={"方案二月度电量(MWh)": "交易量（MWh）", "时段": "时段（点）"},
                        color_discrete_sequence=["#e53e3e"],  # 柔和红
                        height=350
                    )
                    fig.update_layout(
                        plot_bgcolor="white",
                        xaxis_showgrid=False,
                        yaxis_showgrid=True,
                        yaxis_gridcolor="#f0f0f0",
                        font=dict(family="Arial", size=11),
                        title_font=dict(size=13, weight="bold"),
                        margin=dict(l=10, r=10, t=30, b=10)
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    # 准备数据
                    merged_data = arbitrage_df[["时段", "方案二月度电量(MWh)"]].copy()
                    if len(base_df) >= 24:
                        merged_data["现货价格"] = base_df["现货价格(元/MWh)"].head(24).values
                        merged_data["中长期价格"] = base_df["中长期价格(元/MWh)"].head(24).values
                    else:
                        merged_data["现货价格"] = 0.0
                        merged_data["中长期价格"] = 0.0

                    import plotly.graph_objects as go
                    fig = go.Figure()

                    # 交易量柱状图（左轴）
                    fig.add_trace(go.Bar(
                        x=merged_data["时段"],
                        y=merged_data["方案二月度电量(MWh)"],
                        name="方案二交易量",
                        yaxis="y1",
                        marker_color="#e53e3e",  # 柔和红
                        opacity=0.8,
                        hovertemplate="时段：%{x}点<br>交易量：%{y:.2f} MWh<extra></extra>"
                    ))

                    # 现货价格折线（右轴）
                    fig.add_trace(go.Scatter(
                        x=merged_data["时段"],
                        y=merged_data["现货价格"],
                        name="现货价格",
                        yaxis="y2",
                        mode="lines+markers",
                        line=dict(color="#9f7aea", width=2),
                        marker=dict(size=4),
                        hovertemplate="时段：%{x}点<br>现货价格：%{y:.2f} 元/MWh<extra></extra>"
                    ))

                    # 中长期价格折线（右轴）
                    fig.add_trace(go.Scatter(
                        x=merged_data["时段"],
                        y=merged_data["中长期价格"],
                        name="中长期价格",
                        yaxis="y2",
                        mode="lines+markers",
                        line=dict(color="#38b2ac", width=2),
                        marker=dict(size=4),
                        hovertemplate="时段：%{x}点<br>中长期价格：%{y:.2f} 元/MWh<extra></extra>"
                    ))

                    # 视觉优化（和方案一保持风格统一）
                    fig.update_layout(
                        title=f"{view_month}月 方案二交易量与价格对比",
                        title_font=dict(size=13, weight="bold", family="Arial"),
                        title_x=0.5,
                        plot_bgcolor="white",
                        paper_bgcolor="white",
                        yaxis1=dict(
                            title="交易量（MWh）",
                            title_font=dict(color="#e53e3e"),
                            tickfont=dict(color="#e53e3e"),
                            gridcolor="#f0f0f0"
                        ),
                        yaxis2=dict(
                            title="价格（元/MWh）",
                            title_font=dict(color="#9f7aea"),
                            tickfont=dict(color="#9f7aea"),
                            overlaying="y",
                            side="right",
                            gridcolor="rgba(0,0,0,0)"
                        ),
                        legend=dict(
                            orientation="h",
                            yanchor="bottom",
                            y=-0.2,
                            xanchor="center",
                            x=0.5
                        ),
                        margin=dict(l=20, r=20, t=30, b=60),
                        xaxis=dict(
                            title="时段（点）",
                            tickmode="array",
                            tickvals=merged_data["时段"],
                            gridcolor="#f0f0f0"
                        )
                    )

                    st.plotly_chart(fig, use_container_width=True)

        except Exception as e:
            st.warning(f"📊 方案二图表生成失败：{str(e)}（不影响数据导出）")
        
        # 3. 双方案日分解展示（四列数据）
        st.subheader(f"3. {view_month}月双方案日分解电量（四列数据）")
        decompose_df = decompose_double_scheme(
            st.session_state.trade_power_typical[view_month],
            st.session_state.trade_power_arbitrage[view_month],
            st.session_state.current_year,
            view_month
        )
        decompose_df = decompose_df.fillna(0.0)
        # 显示四列核心数据
        display_df = decompose_df[["时段", "方案一月度电量(MWh)", "方案一日分解电量(MWh)", "方案二月度电量(MWh)", "方案二日分解电量(MWh)"]].copy()
        st.dataframe(display_df, use_container_width=True, hide_index=True)
        st.info(f"""
        日分解说明：
        - 日分解电量 = 月度电量 ÷ {view_month}月天数（{get_days_in_month(st.session_state.current_year, view_month)}天）
        - 方案一/二月度总电量保持一致，日分解电量同步匹配
        """)
else:
    if st.session_state.calculated and not st.session_state.trade_power_typical:
        st.warning("⚠️ 无有效方案数据，请重新生成方案")

# -------------------------- 新增：第二步+第三步（比例调整+分时段微调） --------------------------
st.divider()
# 生成独立唯一前缀，避免Key冲突
unique_prefix_ratio_tune = str(uuid.uuid4())[:8]

# 初始化月份切换标记（避免重复初始化）
if f"{unique_prefix_ratio_tune}_month_changed" not in st.session_state:
    st.session_state[f"{unique_prefix_ratio_tune}_month_changed"] = False

# -------------------------- 功能1：月度方案整体比例调整（保持时段占比） --------------------------
st.write("### 📊 月度方案总量比例调整（保持时段占比）")

# 1. 选择调整参数（移除强制rerun的on_change）
col_adjust_1, col_adjust_2, col_adjust_3 = st.columns([2, 2, 1.5])
with col_adjust_1:
    adjust_month = st.selectbox(
        "选择调整月份", 
        range(1, 13), 
        key=f"{unique_prefix_ratio_tune}_ratio_month"
        # 移除on_change=on_month_change，避免强制rerun
    )
with col_adjust_2:
    adjust_scheme = st.selectbox(
        "选择调整方案", 
        ["方案一", "方案二"], 
        key=f"{unique_prefix_ratio_tune}_ratio_scheme"
        # 移除on_change=on_month_change
    )
with col_adjust_3:
    adjust_ratio = st.number_input(
        "调整比例", 
        min_value=0.1, max_value=2.0, value=1.0, step=0.01,
        key=f"{unique_prefix_ratio_tune}_ratio_value",
        help="0.9=90%（缩量）、1.0=不变、1.1=110%（增量）"
    )

# 2. 显示当前数据（强制从session_state读取最新数据，基于选中的adjust_month）
# 初始化当前月份的方案数据（如果不存在）
if adjust_month not in st.session_state.scheme_power_data:
    st.session_state.scheme_power_data[adjust_month] = {
        "方案一": {"periods": {}, "base_total": 0.0},
        "方案二": {"periods": {}, "base_total": 0.0}
    }

current_scheme_data = st.session_state.scheme_power_data[adjust_month][adjust_scheme]
current_periods = current_scheme_data["periods"].copy()  # 强制复制避免引用问题
current_base_total = current_scheme_data["base_total"]

# 兼容现有方案数据（如果scheme_power_data为空，从trade_power_typical/arbitrage读取）
if not current_periods and st.session_state.calculated:
    if adjust_scheme == "方案一" and adjust_month in st.session_state.trade_power_typical:
        # 基于选中的adjust_month读取对应数据
        current_periods = st.session_state.trade_power_typical[adjust_month].set_index("时段")["方案一月度电量(MWh)"].to_dict()
        current_base_total = sum(current_periods.values())
        # 同步到scheme_power_data（基于选中的adjust_month）
        st.session_state.scheme_power_data[adjust_month][adjust_scheme] = {
            "periods": current_periods,
            "base_total": current_base_total
        }
    elif adjust_scheme == "方案二" and adjust_month in st.session_state.trade_power_arbitrage:
        current_periods = st.session_state.trade_power_arbitrage[adjust_month].set_index("时段")["方案二月度电量(MWh)"].to_dict()
        current_base_total = sum(current_periods.values())
        # 同步到scheme_power_data（基于选中的adjust_month）
        st.session_state.scheme_power_data[adjust_month][adjust_scheme] = {
            "periods": current_periods,
            "base_total": current_base_total
        }

current_actual_total = sum(current_periods.values()) if current_periods else 0.0

col_ori_1, col_ori_2 = st.columns(2)
with col_ori_1:
    st.write(f"**{adjust_month}月-{adjust_scheme}**")
    st.write(f"当前基准总量：{current_base_total:.2f} MWh")
    st.write(f"当前实际总量：{current_actual_total:.2f} MWh")
with col_ori_2:
    if current_periods:
        st.write("当前时段电量分布：")
        st.dataframe(
            pd.DataFrame(list(current_periods.items()), columns=["时段", "电量(MWh)"]),
            hide_index=True, use_container_width=True
        )
    else:
        st.warning("该方案暂无时段电量数据，请先生成方案！")

# 3. 执行比例调整（移除强制rerun，改用状态更新）
if st.button(f"✅ 执行{adjust_month}月-{adjust_scheme}比例调整", key=f"{unique_prefix_ratio_tune}_ratio_execute"):
    if not current_periods:
        st.error("调整失败：无基础时段电量数据！")
    else:
        # 步骤1：计算新基准总量（原基准总量×比例，无基准则用实际总量）
        original_base = current_base_total if current_base_total > 0 else current_actual_total
        new_base_total = round(original_base * adjust_ratio, 2)
        
        # 步骤2：按比例缩放各时段电量（保持占比）
        new_periods = {
            period: round(power * adjust_ratio, 2)
            for period, power in current_periods.items()
        }
        
        # 步骤3：更新Session State（基于选中的adjust_month，保证数据不串）
        st.session_state.scheme_power_data[adjust_month][adjust_scheme] = {
            "periods": new_periods,
            "base_total": new_base_total
        }
        # 同步到现有方案数据（保证其他模块能读取到调整后的数据）
        if adjust_scheme == "方案一" and adjust_month in st.session_state.trade_power_typical:
            st.session_state.trade_power_typical[adjust_month]["方案一月度电量(MWh)"] = st.session_state.trade_power_typical[adjust_month]["时段"].map(new_periods)
        elif adjust_scheme == "方案二" and adjust_month in st.session_state.trade_power_arbitrage:
            st.session_state.trade_power_arbitrage[adjust_month]["方案二月度电量(MWh)"] = st.session_state.trade_power_arbitrage[adjust_month]["时段"].map(new_periods)
        
        # 提示结果（移除强制rerun，改用状态刷新）
        st.success(f"""
            比例调整完成！
            基准总量：{original_base:.2f} → {new_base_total:.2f} MWh（比例：{adjust_ratio}）
            各时段电量已按比例缩放，占比保持不变
        """)
        st.write("调整后时段电量：")
        st.dataframe(
            pd.DataFrame(list(new_periods.items()), columns=["时段", "电量(MWh)"]),
            hide_index=True, use_container_width=True
        )

# -------------------------- 功能2：分时段电量微调（自动分摊差额，总量锁定） --------------------------
st.divider()
st.write("### 🛠️ 分时段电量微调（总量锁定为基准值，差额自动分摊）")

# 1. 选择微调参数（移除强制rerun的on_change）
col_tune_1, col_tune_2 = st.columns([2, 2])
with col_tune_1:
    tune_month = st.selectbox(
        "选择微调月份", 
        range(1, 13), 
        key=f"{unique_prefix_ratio_tune}_tune_month"
        # 移除on_change=on_month_change
    )
with col_tune_2:
    tune_scheme = st.selectbox(
        "选择微调方案", 
        ["方案一", "方案二"], 
        key=f"{unique_prefix_ratio_tune}_tune_scheme"
        # 移除on_change=on_month_change
    )

# 初始化微调月份的方案数据（如果不存在）
if tune_month not in st.session_state.scheme_power_data:
    st.session_state.scheme_power_data[tune_month] = {
        "方案一": {"periods": {}, "base_total": 0.0},
        "方案二": {"periods": {}, "base_total": 0.0}
    }

# 获取微调数据（基于选中的tune_month，确保数据不串）
tune_scheme_data = st.session_state.scheme_power_data[tune_month][tune_scheme]
tune_periods = tune_scheme_data["periods"].copy()  # 强制复制避免引用问题
tune_base_total = tune_scheme_data["base_total"]

# 首次微调时，从现有方案数据初始化（基于选中的tune_month）
if not tune_periods and st.session_state.calculated:
    if tune_scheme == "方案一" and tune_month in st.session_state.trade_power_typical:
        tune_periods = st.session_state.trade_power_typical[tune_month].set_index("时段")["方案一月度电量(MWh)"].to_dict()
        tune_base_total = sum(tune_periods.values())
        st.session_state.scheme_power_data[tune_month][tune_scheme] = {
            "periods": tune_periods,
            "base_total": tune_base_total
        }
    elif tune_scheme == "方案二" and tune_month in st.session_state.trade_power_arbitrage:
        tune_periods = st.session_state.trade_power_arbitrage[tune_month].set_index("时段")["方案二月度电量(MWh)"].to_dict()
        tune_base_total = sum(tune_periods.values())
        st.session_state.scheme_power_data[tune_month][tune_scheme] = {
            "periods": tune_periods,
            "base_total": tune_base_total
        }

if not tune_periods:
    st.warning("该方案暂无时段电量数据，请先生成/调整方案！")
else:
    if tune_base_total <= 0:
        st.warning("请先执行「比例调整」设置基准总量！")
    else:
        # 显示基准总量（基于选中的tune_month）
        st.info(f"🔒 锁定基准总量：{tune_base_total:.2f} MWh（修改时段后自动分摊差额）")
        
        # 2. 选择要修改的时段+输入新值
        col_tune_3, col_tune_4 = st.columns([2, 2])
        with col_tune_3:
            target_period = st.selectbox("选择要修改的时段", list(tune_periods.keys()), key=f"{unique_prefix_ratio_tune}_tune_period")
        with col_tune_4:
            new_power = st.number_input(
                f"{target_period} 新电量(MWh)",
                min_value=0.0, value=tune_periods[target_period], step=0.1,
                key=f"{unique_prefix_ratio_tune}_tune_power"
            )
        
        # 3. 执行微调（核心：自动分摊差额，基于选中的tune_month）
        if st.button(f"✅ 执行{target_period}电量微调", key=f"{unique_prefix_ratio_tune}_tune_execute"):
            # 步骤1：计算差额（目标值 - 原值）
            original_power = tune_periods[target_period]
            diff = round(new_power - original_power, 2)
            
            if diff == 0:
                st.info("无差额：新值与原值一致！")
            else:
                # 步骤2：获取其他时段（排除当前修改的时段）
                other_periods = {p: v for p, v in tune_periods.items() if p != target_period}
                other_total = sum(other_periods.values())
                
                if not other_periods:
                    st.error("无法分摊：仅单个时段，需手动调整总量！")
                else:
                    # 步骤3：按占比分摊差额（保证其他时段总量 ±diff，占比不变）
                    new_other_periods = {}
                    for p, v in other_periods.items():
                        # 分摊系数 = 该时段占其他时段总量的比例
                        ratio = v / other_total
                        # 该时段需调整的量 = -diff × 比例（反向分摊，抵消差额）
                        p_diff = round(-diff * ratio, 2)
                        new_v = round(v + p_diff, 2)
                        # 防止负数
                        new_other_periods[p] = max(new_v, 0.01)
                    
                    # 步骤4：更新所有时段电量（基于选中的tune_month）
                    updated_periods = {**new_other_periods, target_period: new_power}
                    # 最终校验：总量强制等于基准值（解决浮点误差）
                    final_total = sum(updated_periods.values())
                    total_diff = round(tune_base_total - final_total, 2)
                    if abs(total_diff) > 0.01:
                        # 误差分摊到第一个其他时段
                        first_p = list(new_other_periods.keys())[0]
                        updated_periods[first_p] = round(updated_periods[first_p] + total_diff, 2)
                    
                    # 步骤5：更新Session State（基于选中的tune_month，保证数据不串）
                    st.session_state.scheme_power_data[tune_month][tune_scheme]["periods"] = updated_periods
                    # 同步到trade_power_typical/arbitrage，保证其他模块联动
                    if tune_scheme == "方案一" and tune_month in st.session_state.trade_power_typical:
                        st.session_state.trade_power_typical[tune_month]["方案一月度电量(MWh)"] = st.session_state.trade_power_typical[tune_month]["时段"].map(updated_periods)
                    elif tune_scheme == "方案二" and tune_month in st.session_state.trade_power_arbitrage:
                        st.session_state.trade_power_arbitrage[tune_month]["方案二月度电量(MWh)"] = st.session_state.trade_power_arbitrage[tune_month]["时段"].map(updated_periods)
                    
                    # 提示结果（移除强制rerun）
                    st.success(f"""
                        微调完成！
                        {target_period}：{original_power:.2f} → {new_power:.2f} MWh（差额：{diff:.2f}）
                        差额已自动分摊到其他时段，总量锁定为 {tune_base_total:.2f} MWh
                    """)
                    st.write("微调后所有时段电量：")
                    st.dataframe(
                        pd.DataFrame(list(updated_periods.items()), columns=["时段", "电量(MWh)"]),
                        hide_index=True, use_container_width=True
                    )

# -------------------------- 新增：收益计算功能（缩进+逻辑修复）--------------------------
st.divider()
st.header("💰 双方案收益计算（实时同步电量调整结果）")

# 仅当方案生成成功且有有效数据时计算收益（最外层if）
if st.session_state.calculated and st.session_state.trade_power_typical and st.session_state.trade_power_arbitrage:
    # 过滤有完整收益数据的月份（需包含电量+价格数据）
    valid_profit_months = []
    for month in st.session_state.selected_months:
        # 校验方案数据（电量）
        has_plan1 = (month in st.session_state.trade_power_typical 
                    and not st.session_state.trade_power_typical[month].empty
                    and "方案一月度电量(MWh)" in st.session_state.trade_power_typical[month].columns)
        has_plan2 = (month in st.session_state.trade_power_arbitrage 
                    and not st.session_state.trade_power_arbitrage[month].empty
                    and "方案二月度电量(MWh)" in st.session_state.trade_power_arbitrage[month].columns)
        # 校验价格数据（现货+中长期，至少有一个价格不为0）
        has_price = (month in st.session_state.monthly_data 
                    and not st.session_state.monthly_data[month].empty
                    and "现货价格(元/MWh)" in st.session_state.monthly_data[month].columns
                    and "中长期价格(元/MWh)" in st.session_state.monthly_data[month].columns
                    and (st.session_state.monthly_data[month]["现货价格(元/MWh)"].sum() > 0 
                         or st.session_state.monthly_data[month]["中长期价格(元/MWh)"].sum() > 0))
        
        if has_plan1 and has_plan2 and has_price:
            valid_profit_months.append(month)
    
    # 第一层嵌套if：有有效收益月份
    if valid_profit_months:
        # 选择收益计算的月份（默认全选有效月份）
        profit_months = st.multiselect(
            "选择需要计算收益的月份",
            options=valid_profit_months,
            default=valid_profit_months,
            key="profit_month_select",
            format_func=lambda x: f"{x}月"
        )
        
        # 第二层嵌套if：选择了计算月份
        if profit_months:
            # 初始化年度收益汇总
            annual_profit_plan1 = 0.0  # 方案一年度总收益
            annual_profit_plan2 = 0.0  # 方案二年度总收益
            monthly_profit_list = []   # 月度收益明细
            
            # 循环计算每个选中月份的收益
            for month in profit_months:
                plan1_df = st.session_state.trade_power_typical[month].copy()
                plan2_df = st.session_state.trade_power_arbitrage[month].copy()
                price_df = st.session_state.monthly_data[month].copy()
                
                # 取前24时段数据（确保电量和价格一一对应）
                plan1_power = plan1_df["方案一月度电量(MWh)"].head(24).values
                plan2_power = plan2_df["方案二月度电量(MWh)"].head(24).values
                spot_price = price_df["现货价格(元/MWh)"].head(24).values
                mid_long_price = price_df["中长期价格(元/MWh)"].head(24).values
                
                # 计算时段收益（电量×价格，价格优先取现货，现货为0则取中长期）
                plan1_hourly_profit = []
                plan2_hourly_profit = []
                for i in range(24):
                    # 选择有效价格（现货>0用现货，否则用中长期）
                    use_price = spot_price[i] if spot_price[i] > 0 else mid_long_price[i]
                    use_price = max(use_price, 0)  # 避免负价格导致收益异常
                    
                    # 计算单个时段收益
                    p1_profit = round(plan1_power[i] * use_price, 2)
                    p2_profit = round(plan2_power[i] * use_price, 2)
                    
                    plan1_hourly_profit.append(p1_profit)
                    plan2_hourly_profit.append(p2_profit)
                
                # 计算月度总收益
                monthly_profit1 = sum(plan1_hourly_profit)
                monthly_profit2 = sum(plan2_hourly_profit)
                
                # 累加年度收益
                annual_profit_plan1 += monthly_profit1
                annual_profit_plan2 += monthly_profit2
                
                # 保存月度明细（含更优方案标记）
                if monthly_profit1 > monthly_profit2:
                    better_plan = f"**<span style='color: #22c55e'>方案一</span>**"
                elif monthly_profit2 > monthly_profit1:
                    better_plan = f"**<span style='color: #ef4444'>方案二</span>**"
                else:
                    better_plan = f"**<span style='color: #64748b'>持平</span>**"
                
                monthly_profit_list.append({
                    "月份": f"{month}月",
                    "方案一收益（元）": monthly_profit1,
                    "方案二收益（元）": monthly_profit2,
                    "收益差值（方案二-方案一）": round(monthly_profit2 - monthly_profit1, 2),
                    "更优方案": better_plan
                })
            
            # 生成月度数据DataFrame
            profit_detail_df = pd.DataFrame(monthly_profit_list)
            
            # 追加「年度汇总行」
            annual_better_plan = (
                f"**<span style='color: #22c55e'>方案一</span>**" if annual_profit_plan1 > annual_profit_plan2
                else f"**<span style='color: #ef4444'>方案二</span>**" if annual_profit_plan2 > annual_profit_plan1
                else f"**<span style='color: #64748b'>持平</span>**"
            )
            
            annual_summary = pd.DataFrame([{
                "月份": "年度汇总",
                "方案一收益（元）": annual_profit_plan1,
                "方案二收益（元）": annual_profit_plan2,
                "收益差值（方案二-方案一）": round(annual_profit_plan2 - annual_profit_plan1, 2),
                "更优方案": annual_better_plan
            }])
            
            # 合并月度数据和汇总行
            profit_detail_df = pd.concat([profit_detail_df, annual_summary], ignore_index=True)
            
            # 显示分月收益对比表格
            st.subheader("📋 分月收益对比（含月度更优方案）")
            st.dataframe(
                profit_detail_df,
                use_container_width=True,
                column_config={
                    "月份": st.column_config.TextColumn("月份", width="small"),
                    "方案一收益（元）": st.column_config.NumberColumn("方案一收益（元）", format="¥%.2f"),
                    "方案二收益（元）": st.column_config.NumberColumn("方案二收益（元）", format="¥%.2f"),
                    "收益差值（方案二-方案一）": st.column_config.NumberColumn(
                        "收益差值（方案二-方案一）",
                        format="¥%.2f",
                        help="正值=方案二更优，负值=方案一更优"
                    ),
                    "更优方案": st.column_config.TextColumn(
                        "更优方案",
                        help="当月收益更高的方案（绿色=方案一，红色=方案二，灰色=持平）"
                    )
                }
            )
            
            # 显示年度收益汇总（卡片式）
            st.subheader("📊 年度收益汇总")
            col_p1, col_p2, col_diff = st.columns(3, gap="large")
            
            with col_p1:
                st.metric(
                    label="方案一年度总收益",
                    value=f"¥{round(annual_profit_plan1, 2):,.2f}",
                    delta=None,
                    help="基于典型曲线电量计算"
                )
            
            with col_p2:
                st.metric(
                    label="方案二年度总收益",
                    value=f"¥{round(annual_profit_plan2, 2):,.2f}",
                    delta=None,
                    help="基于套利/直线曲线电量计算"
                )
            
            with col_diff:
                profit_diff = round(annual_profit_plan2 - annual_profit_plan1, 2)
                delta_color = "normal" if profit_diff == 0 else ("inverse" if profit_diff < 0 else "off")
                st.metric(
                    label="方案二相对方案一收益差",
                    value=f"¥{profit_diff:,.2f}",
                    delta=f"{profit_diff/annual_profit_plan1*100:.2f}%" if annual_profit_plan1 != 0 else "无参考",
                    delta_color=delta_color,
                    help="正值=方案二更优，负值=方案一更优"
                )
            
            # 收益计算说明（放在最内层if里，只有选择了月份才显示）
            st.caption("""
            📌 收益计算规则：
            1. 价格优先级：优先使用「现货价格」，现货价格为0时使用「中长期价格」；
            2. 时段收益=时段电量 × 对应价格（保留2位小数）；
            3. 实时同步：手动调整电量后，收益会自动重新计算（无需额外操作）；
            4. 数据安全：收益计算不修改任何原始数据，仅基于现有方案和价格数据统计。
            """)
        
        # 对应第二层if：未选择计算月份
        else:
            st.info("ℹ️ 请选择需要计算收益的月份")
    
    # 对应第一层if：无有效收益月份
    else:
        st.info("ℹ️ 暂无有效收益计算数据，请确保：1. 生成了年度方案 2. 模板中填写了现货/中长期价格（非0） 3. 选中月份有完整数据")

# 对应最外层if：未生成方案
else:
    st.warning("⚠️ 请先生成年度方案后，再计算收益")

# 页脚
st.divider()
st.caption(f"© {st.session_state.current_year} 新能源电厂年度方案设计系统 | 双方案（典型/套利/直线）+ 四列日分解数据 | 总量一致")
