import streamlit as st
import pandas as pd
import chardet
import warnings
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import tempfile
import os
from io import BytesIO

# ========== 全局配置与警告屏蔽 ==========
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
warnings.filterwarnings("ignore", category=UserWarning, module="pandas")

# Streamlit 页面配置
st.set_page_config(
    page_title="河南功率预测 - 买入/卖出模板生成工具",
    page_icon="⚡",
    layout="wide"
)

# 全局配置项（可在界面修改）
DEFAULT_TIME_FORMAT = "%Y-%m-%d_%H:%M:%S"
DEFAULT_DECIMAL_PLACES = 3

# ========== 工具函数：保持原有核心逻辑，适配 Streamlit 上传文件 ==========
def detect_file_encoding(file_bytes):
    """检测文件编码（适配字节流）"""
    result = chardet.detect(file_bytes[:10000])
    return result['encoding'] or 'utf-8'

def read_power_txt_from_bytes(file_bytes):
    """从上传的 txt 字节流中读取功率数据"""
    encoding = detect_file_encoding(file_bytes)
    raw_lines = []
    lines = file_bytes.decode(encoding, errors="ignore").split("\n")
    
    for line in lines:
        line = line.strip()
        # 过滤空行、HTML标签行、表头行（times prepower）
        if not line or line.startswith("<!") or line.lower() == "times prepower":
            continue
        parts = line.split()
        if len(parts) >= 2:
            raw_lines.append([parts[0], parts[1]])
    
    if not raw_lines:
        raise Exception("❌ 过滤后无有效功率数据！请检查 txt 文件内容")
    power_df = pd.DataFrame(raw_lines, columns=["times", "prepower"])
    return power_df

def build_full_time_grid(power_df, time_col, time_format, decimal_places):
    """构建全量时段网格（保留核心逻辑，适配传入配置）"""
    power_df = power_df.copy(deep=True)
    
    # 时间列转换
    power_df[time_col] = pd.to_datetime(power_df[time_col], format=time_format, errors="coerce")
    power_df = power_df.dropna(subset=[time_col])
    if power_df.empty:
        raise Exception(f"❌ 时间列转换后无有效数据！请检查时间格式是否为 {time_format}")
    
    # 构建全量日期+小时网格
    all_dates = power_df[time_col].dt.date.unique()
    all_hours = list(range(24))
    full_grid = pd.MultiIndex.from_product([all_dates, all_hours], names=["日期", "小时"]).to_frame(index=False)
    
    # 计算小时级平均发电量
    power_df["prepower"] = pd.to_numeric(power_df["prepower"], errors="coerce").fillna(0)
    power_df["日期"] = power_df[time_col].dt.date
    power_df["小时"] = power_df[time_col].dt.hour
    hourly_power = power_df.groupby(["日期", "小时"])["prepower"].mean().reset_index()
    hourly_power.columns = ["日期", "小时", "发电量"]
    
    # 合并全量网格，填充0值
    full_power = pd.merge(full_grid, hourly_power, on=["日期", "小时"], how="left")
    full_power["发电量"] = full_power["发电量"].fillna(0)
    full_power["0.2倍发电量"] = round(full_power["发电量"] * 0.2, decimal_places)
    
    return full_power, all_dates, all_hours

def reshape_to_wide_table(full_power, all_dates, decimal_places):
    """重塑为时段行、日期列宽表"""
    power_only = full_power[["日期", "小时", "发电量"]].copy()
    power_02x = full_power[["日期", "小时", "0.2倍发电量"]].copy()
    
    # 透视成宽表
    power_wide = power_only.pivot(index="小时", columns="日期", values="发电量")
    power_02x_wide = power_02x.pivot(index="小时", columns="日期", values="0.2倍发电量")
    
    # 重命名列
    power_wide.columns = [f"{dt}_发电量" for dt in power_wide.columns]
    power_02x_wide.columns = [f"{dt}_0.2倍发电量" for dt in power_02x_wide.columns]
    
    # 合并并排序列
    wide_table = pd.concat([power_wide, power_02x_wide], axis=1)
    wide_table = wide_table.reindex(columns=sorted(wide_table.columns, key=lambda x: (x.split("_")[0], x.split("_")[1])))
    
    # 重命名行（H1-H24）
    wide_table.index.name = "时段"
    wide_table.index = [f"H{h+1}" for h in wide_table.index]
    
    return wide_table

def generate_excel_file(template_file_bytes, full_power, is_buy_template, decimal_places):
    """生成买入/卖出 Excel 模板（返回字节流，用于下载）"""
    # 加载上传的模板文件
    wb = load_workbook(BytesIO(template_file_bytes))
    target_sheet = wb.sheetnames[0]
    ws = wb[target_sheet]
    
    # 读取模板数据
    template_df = pd.read_excel(BytesIO(template_file_bytes), sheet_name=target_sheet)
    if "交易代码" not in template_df.columns:
        raise Exception("❌ 原模板中未找到'交易代码'列！无法匹配时段")
    
    # 解析交易代码
    def parse_trade_code(code):
        if pd.isna(code) or not str(code).startswith("D") or len(str(code)) < 10:
            return None, None
        try:
            date_str = str(code)[1:9]
            hour_part = str(code).split("H")[-1]
            if not hour_part.isdigit():
                return None, None
            date = datetime.strptime(date_str, "%Y%m%d").date()
            hour = int(hour_part) - 1
            if hour < 0 or hour > 23:
                return None, None
            return date, hour
        except:
            return None, None
    
    template_df[["日期", "小时"]] = template_df["交易代码"].apply(
        lambda x: pd.Series(parse_trade_code(x))
    )
    
    # 关联功率数据，计算可交易电量
    merged_df = pd.merge(
        template_df,
        full_power[["日期", "小时", "0.2倍发电量"]],
        on=["日期", "小时"],
        how="left"
    )
    
    # 处理分时限额
    merged_df["分时限额"] = pd.to_numeric(merged_df["分时限额"], errors="coerce").fillna(0)
    merged_df["分时限额"] = round(merged_df["分时限额"], decimal_places)
    merged_df["0.2倍发电量"] = merged_df["0.2倍发电量"].fillna(0)
    merged_df["可交易电量"] = round(merged_df["0.2倍发电量"] - merged_df["分时限额"], decimal_places)
    
    # 清空原模板数据行
    max_row = ws.max_row
    max_col = ws.max_column
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col, value=None)
    
    # 卖出模板专属处理：修改E、I列表头，填充E列
    if not is_buy_template:
        ws.cell(row=1, column=5, value="卖出电量")
        ws.cell(row=1, column=9, value="卖出电价")
        
        for idx, (_, row_data) in enumerate(merged_df.iterrows(), 2):
            trade_power = row_data["可交易电量"]
            if trade_power > 0:
                ws.cell(row=idx, column=5, value=trade_power)
            else:
                ws.cell(row=idx, column=5, value=0.0)
    
    # 买入模板专属处理：填充E列，标黄
    else:
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for idx, (_, row_data) in enumerate(merged_df.iterrows(), 2):
            trade_power = row_data["可交易电量"]
            if trade_power < 0:
                ws.cell(row=idx, column=5, value=round(abs(trade_power), decimal_places))
                cell = ws.cell(row=idx, column=5)
                cell.fill = yellow_fill
            else:
                ws.cell(row=idx, column=5, value=0.0)
    
    # 保存为字节流（用于下载）
    output_bytes = BytesIO()
    wb.save(output_bytes)
    output_bytes.seek(0)
    wb.close()
    
    return output_bytes

# ========== Streamlit 可视化界面构建 ==========
def main():
    # 页面标题
    st.title("⚡ 河南功率预测 - 买入/卖出模板在线生成工具")
    st.divider()
    
    # 侧边栏配置
    with st.sidebar:
        st.header("⚙️ 配置项")
        time_format = st.text_input("时间格式", value=DEFAULT_TIME_FORMAT, help="txt 文件中的时间格式，默认：%Y-%m-%d_%H:%M:%S")
        decimal_places = st.number_input("保留小数位数", min_value=1, max_value=6, value=DEFAULT_DECIMAL_PLACES, step=1)
        st.info("""
        使用说明：
        1. 上传功率预测 txt 文件
        2. 上传电量电价导入模板 Excel
        3. 点击「开始处理」
        4. 下载生成的结果文件
        """)
    
    # 主界面：文件上传
    col1, col2 = st.columns(2)
    with col1:
        txt_file = st.file_uploader("📄 上传功率预测 TXT 文件", type=["txt"])
    with col2:
        excel_template_file = st.file_uploader("📊 上传 Excel 模板文件", type=["xlsx"])
    
    st.divider()
    
    # 处理按钮与核心逻辑
    if st.button("🚀 开始处理", type="primary", disabled=(not txt_file or not excel_template_file)):
        try:
            with st.spinner("正在处理数据，请稍候..."):
                # 1. 读取并处理 txt 功率文件
                txt_bytes = txt_file.read()
                power_df = read_power_txt_from_bytes(txt_bytes)
                
                # 2. 构建全量时段数据
                full_power, all_dates, all_hours = build_full_time_grid(
                    power_df,
                    time_col="times",
                    time_format=time_format,
                    decimal_places=decimal_places
                )
                
                # 3. 生成发电量宽表
                wide_table = reshape_to_wide_table(full_power, all_dates, decimal_places)
                
                # 4. 生成买入/卖出模板 Excel 字节流
                excel_template_bytes = excel_template_file.read()
                buy_excel_bytes = generate_excel_file(
                    excel_template_bytes,
                    full_power,
                    is_buy_template=True,
                    decimal_places=decimal_places
                )
                sell_excel_bytes = generate_excel_file(
                    excel_template_bytes,
                    full_power,
                    is_buy_template=False,
                    decimal_places=decimal_places
                )
                
                # 5. 生成宽表 Excel 字节流
                wide_excel_bytes = BytesIO()
                wide_table.to_excel(wide_excel_bytes, index=True)
                wide_excel_bytes.seek(0)
            
            # 处理完成：展示结果与下载
            st.success("✅ 数据处理完成！")
            st.divider()
            
            # 数据预览
            st.subheader("📄 数据预览")
            tab1, tab2, tab3 = st.tabs(["功率数据", "全量时段数据", "发电量宽表"])
            with tab1:
                st.dataframe(power_df.head(10), use_container_width=True)
            with tab2:
                st.dataframe(full_power.head(10), use_container_width=True)
            with tab3:
                st.dataframe(wide_table.head(10), use_container_width=True)
            
            st.divider()
            
            # 下载区域
            st.subheader("📥 结果下载")
            col_download1, col_download2, col_download3 = st.columns(3)
            with col_download1:
                st.download_button(
                    label="下载 买入模板",
                    data=buy_excel_bytes,
                    file_name="电量电价导入模板_买入.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with col_download2:
                st.download_button(
                    label="下载 卖出模板",
                    data=sell_excel_bytes,
                    file_name="电量电价导入模板_卖出.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with col_download3:
                st.download_button(
                    label="下载 发电量宽表",
                    data=wide_excel_bytes,
                    file_name="发电量_日期列_时段行.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        except Exception as e:
            st.error(f"❌ 处理失败：{str(e)}")
            st.exception(e)

if __name__ == "__main__":
    main()
