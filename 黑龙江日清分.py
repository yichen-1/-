import streamlit as st
import pandas as pd
import re
from datetime import datetime
import warnings
import pdfplumber
from io import BytesIO
import sys
import os
from openpyxl.styles import PatternFill

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.stylesheet")

# ---------------------- 核心配置（新增阻塞/价差费用专属配置） ----------------------
REDUNDANT_KEYWORDS = [
    "内部使用", "CONFIDENTIAL", "草稿", "现货试结算期间", "日清分单据",
    "公司名称", "编号：", "单位：", "清分日期", "合计电量", "合计电费",
    "电能量电费", "审批：", "审核：", "编制：", "加盖电子签章",
    "ylxxhfd", "yxxchfd", "依兰县协合风力发电有限公司", "依", "依兰", "协合",
    "县", "风力发电", "有限公司", "司"
]
TRADE_CODE_MAP = {
    "0101010101": "优先发电交易",
    "0101020101": "电网企业代理购电交易", 
    "0101020301": "省内电力直接交易",
    "0101040203": "送上海省间绿色电力交易(电能量)",
    "0101040301": "送辽宁交易",
    "0101040321": "送华北交易", 
    "0101040322": "送山东交易",
    "0101040330": "送浙江交易",
    "0102020101": "省内现货日前交易",
    "0102020301": "省内现货实时交易",
    "0102010101": "省间现货日前交易",
    "0102010201": "省间现货日内交易",
    "0202030001": "中长期合约阻塞费用",  # 重点科目
    "0202030002": "省间省内价差费用",   # 重点科目
    "0101070101": "现货结算价差调整",
    "0101090101": "辅助服务费用分摊",
    "0101100101": "偏差考核费用",
    "0101020201": "省内绿色电力交易(电能量)",
    "0101040202": "送江苏省间绿色电力交易(电能量)",
    "0101040204": "送浙江省间绿色电力交易(电能量)",
    "0101100001": "日融合交易",
    "101010101": "优先发电交易",
    "101020201": "省内绿色电力交易(电能量)",
    "101040202": "送江苏省间绿色电力交易(电能量)",
    "101040204": "送浙江省间绿色电力交易(电能量)",
    "101100001": "日融合交易",
    "101040330": "送浙江交易",
    "101070101": "现货结算价差调整",
    "101090101": "辅助服务费用分摊",
    "101100101": "偏差考核费用"
}
TRADE_KEYWORDS = {
    "优先发电": "优先发电交易",
    "代理购电": "电网企业代理购电交易",
    "直接交易": "省内电力直接交易",
    "绿色电力": "省内绿色电力交易(电能量)",
    "送江苏": "送江苏省间绿色电力交易(电能量)",
    "送浙江": "送浙江交易",
    "送浙江省间": "送浙江省间绿色电力交易(电能量)",
    "送上海": "送上海省间绿色电力交易(电能量)",
    "送辽宁": "送辽宁交易",
    "送华北": "送华北交易",
    "送山东": "送山东交易",
    "日融合": "日融合交易",
    "现货日前": "省内现货日前交易",
    "现货实时": "省内现货实时交易",
    "阻塞费用": "中长期合约阻塞费用",       # 强化关键词
    "中长期合约阻塞": "中长期合约阻塞费用", # 新增变体
    "价差费用": "省间省内价差费用",        # 强化关键词
    "省间省内价差": "省间省内价差费用",     # 新增变体
    "现货结算": "现货结算价差调整",
    "辅助服务": "辅助服务费用分摊",
    "偏差考核": "偏差考核费用"
}
# 调整数据规则：允许阻塞/价差费用的电费为负数且范围更广
DATA_RULES = {
    "电量(兆瓦时)": {"min": -1000, "max": 5000},
    "电价(元/兆瓦时)": {"min": 0, "max": 2000},
    "电费(元)": {"min": -1000000, "max": 10000000}  # 扩大负数范围
}
# 1. 新增晶盛光伏站变体，覆盖“晶盛光伏站”→“晶盛光伏电站”
STATION_CORE_NAMES = [
    "双发A风电场", "双发B风电场", 
    "晶盛光伏电站", "晶盛光伏站",  # 新增晶盛变体
    "晶盛光伏",  # 极简变体
    "双A风场", "双B风场", "双发A", "双发B"
]
STATION_SPLIT_KEYWORDS = ["机组", "机组名称", "双发A", "双发B", "晶盛", "风场", "风电场", "光伏"]
# 2. 新增“光伏站”类型关键词，便于识别补全
STATION_TYPE_KEYWORDS = ["风电场", "风场", "光伏电站", "光伏站", "电站", "场站"]
# 3. 确保“计量量”在最前，优先截断
EXCLUDE_KEYWORDS = ["计量量", "计量电量", "电量", "电价", "电费", "合计", "小计"]
# 仅针对场站名称的数字/冗余过滤
STATION_NUMBER_PATTERN = re.compile(r'\s+\d+\.?\d*')
# 场站名称冗余词正则（匹配“ 计量量”“ 计量电量”）
STATION_REDUNDANT_PATTERN = re.compile(r'\s+(计量量|计量电量|电量|电价|电费)')

# ---------------------- 核心工具函数（重点修复阻塞/价差费用提取） ----------------------
def remove_redundant_text(text):
    if not text:
        return ""
    cleaned = str(text).strip()
    for keyword in REDUNDANT_KEYWORDS:
        if keyword not in ["机组", "电量", "电价", "电费", "晶盛"]:  # 保留晶盛关键词
            cleaned = cleaned.replace(keyword, "")
    single_watermarks = ["依", "兰", "协", "合", "县", "电", "力", "发", "限", "司"]
    for char in single_watermarks:
        cleaned = cleaned.replace(char, "")
    cleaned = re.sub(r'\s+', ' ', cleaned)
    cleaned = re.sub(r'[^\u4e00-\u9fa5a-zA-Z0-9\.\-\:\s]', '', cleaned)
    return cleaned.strip()

def clean_station_name(station_name):
    """核心修复：晶盛光伏站补全+冗余词截断"""
    if not station_name or station_name == "未知场站":
        return "未知场站"
    # 1. 先剔除场站名称后的冗余词（如“晶盛光伏站 计量量”→“晶盛光伏站”）
    cleaned = STATION_REDUNDANT_PATTERN.sub('', station_name)
    # 2. 剔除数字（如“晶盛光伏站 123”→“晶盛光伏站”）
    cleaned = STATION_NUMBER_PATTERN.sub('', cleaned)
    # 3. 标准化名称：补全“晶盛光伏站”→“晶盛光伏电站”
    standard_map = {
        "双发A风电场": "双发A风电场",
        "双发B风电场": "双发B风电场",
        "双A风场": "双发A风电场",
        "双B风场": "双发B风电场",
        "双发A": "双发A风电场",
        "双发B": "双发B风电场",
        "双A": "双发A风电场",
        "双B": "双发B风电场",
        # 新增晶盛光伏站标准化映射
        "晶盛光伏电站": "晶盛光伏电站",
        "晶盛光伏站": "晶盛光伏电站",
        "晶盛光伏": "晶盛光伏电站"
    }
    # 优先匹配标准化名称
    for variant, standard in standard_map.items():
        if variant in cleaned:
            return standard
    # 4. 兜底：补全光伏站→光伏电站
    if "光伏站" in cleaned:
        return cleaned.replace("光伏站", "光伏电站")
    # 5. 保留其他有效场站名
    for type_key in STATION_TYPE_KEYWORDS:
        if type_key in cleaned:
            match = re.search(r'([^，。\n]+' + type_key + ')', cleaned)
            if match:
                return match.group(1).strip()
    return cleaned.strip()

def extract_station_from_text(pdf_text):
    """修复：优先提取晶盛光伏相关名称"""
    clean_text = remove_redundant_text(pdf_text)
    # 1. 优先匹配晶盛光伏相关（确保不遗漏）
    jingsheng_patterns = [
        r'机组[:：\s]*([^，。\n]+晶盛[^，。\n]+光伏电站|[^，。\n]+晶盛[^，。\n]+光伏站)',
        r'机组名称[:：\s]*([^，。\n]+晶盛[^，。\n]+光伏电站|[^，。\n]+晶盛[^，。\n]+光伏站)',
        r'(晶盛光伏电站|晶盛光伏站|晶盛光伏)'
    ]
    for pattern in jingsheng_patterns:
        match = re.search(pattern, clean_text)
        if match:
            raw_name = match.group(1).strip()
            return clean_station_name(raw_name)
    # 2. 匹配双发风电场
    shuangfa_patterns = [
        r'机组[:：\s]*([^，。\n]+双发[^，。\n]+风电场|[^，。\n]+双发[^，。\n]+风场)',
        r'(双发A风电场|双发B风电场|双A风场|双B风场|双发A|双发B)'
    ]
    for pattern in shuangfa_patterns:
        match = re.search(pattern, clean_text)
        if match:
            raw_name = match.group(1).strip()
            return clean_station_name(raw_name)
    # 3. 兜底提取
    for type_key in STATION_TYPE_KEYWORDS:
        match = re.search(r'([^，。\n]+' + type_key + ')', clean_text)
        if match:
            raw_name = match.group(1).strip()
            return clean_station_name(raw_name)
    return "未知场站"

def extract_station_from_filename(file_name):
    """修复：从文件名提取晶盛光伏站"""
    if not file_name:
        return "未知场站"
    # 1. 优先匹配晶盛光伏
    if "晶盛" in file_name and "光伏" in file_name:
        return "晶盛光伏电站"
    # 2. 匹配双发风电场
    if "双发" in file_name or "双A" in file_name or "双B" in file_name:
        shuangfa_match = re.search(r'(双发[AB]风电场|双[AB]风场|双发[AB]|双[AB])', file_name)
        if shuangfa_match:
            return clean_station_name(shuangfa_match.group(1))
    # 3. 兜底
    for type_key in STATION_TYPE_KEYWORDS:
        match = re.search(r'([^_]+' + type_key + ')', file_name)
        if match:
            return clean_station_name(match.group(1))
    return "未知场站"

def safe_convert_to_numeric(value, data_type=""):
    """修复：强化负数处理，适配阻塞/价差费用的负数电费"""
    if value is None or pd.isna(value) or value == '':
        return None
    val_str = remove_redundant_text(value)
    if re.match(r'^\d{9,10}$', val_str):
        return val_str
    if val_str in ['-', '.', '', '—', '——']:
        return None
    try:
        # 修复：保留负号，正确处理负数
        cleaned = re.sub(r'[^\d\-\.]', '', val_str.replace('，', ',').replace('。', '.'))
        if not cleaned or cleaned in ['-', '.', '-.' , '-.']:
            return None
        num = float(cleaned)
        # 修复：阻塞/价差费用不受电价最小值限制
        if data_type in DATA_RULES:
            rule = DATA_RULES[data_type]
            # 电价仅对非阻塞/价差费用做最小值限制
            if data_type == "电价(元/兆瓦时)" and "阻塞" not in data_type and "价差" not in data_type:
                if num < rule["min"] or num > rule["max"]:
                    return None
            else:
                if num < rule["min"] or num > rule["max"]:
                    return None
        return num
    except (ValueError, TypeError):
        return None

def extract_company_info(pdf_text, file_name):
    clean_text = remove_redundant_text(pdf_text)
    company_name = "未知发电公司"
    company_match = re.search(r'公司名称[:：]\s*([^，。\n]+公司)', clean_text)
    if company_match:
        company_name = company_match.group(1).strip()
    else:
        company_match = re.search(r'([^_]+公司|[^_]+发电)', file_name)
        if company_match:
            company_name = company_match.group(1).strip()
    return company_name

def extract_clear_date(pdf_text, file_name):
    raw_text = str(pdf_text).strip()
    date = None
    date_patterns = [
        r'清分日期[:：]\s*(\d{4}-\d{1,2}-\d{1,2}|\d{4}年\d{1,2}月\d{1,2}日)',
        r'结算日期[:：]\s*(\d{4}-\d{1,2}-\d{1,2}|\d{4}年\d{1,2}月\d{1,2}日)',
        r'(\d{4}-\d{1,2}-\d{1,2})'
    ]
    for pattern in date_patterns:
        match = re.search(pattern, raw_text)
        if match:
            date_str = match.group(1).strip()
            if "年" in date_str:
                date_str = date_str.replace("年", "-").replace("月", "-").replace("日", "")
            date = date_str
            break
    if not date:
        date_match = re.search(r'(\d{4}-\d{1,2}-\d{1,2}|\d{8})', file_name)
        if date_match:
            date_str = date_match.group(1)
            if len(date_str) == 8:
                date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
            else:
                date = date_str
    return date

def split_double_station_tables(all_tables, pdf_text, file_name):
    clean_text = remove_redundant_text(pdf_text)
    merged_rows = []
    for table in all_tables:
        if not table:
            continue
        for row in table:
            cleaned_row = [remove_redundant_text(cell) for cell in row]
            if any(cell.strip() != "" for cell in cleaned_row):
                merged_rows.append(cleaned_row)
    
    if not merged_rows:
        text_station = extract_station_from_text(pdf_text)
        return [(text_station, [])] if text_station != "未知场站" else [(extract_station_from_filename(file_name), [])]
    
    station_segments = []
    current_segment = []
    current_station = extract_station_from_text(pdf_text)
    
    for row in merged_rows:
        row_str = ''.join(row).replace(" ", "")
        has_station_key = any(keyword in row_str for keyword in STATION_SPLIT_KEYWORDS)
        
        if has_station_key:
            if current_segment:
                cleaned_station = clean_station_name(current_station)
                station_segments.append((cleaned_station, current_segment))
            row_text = ' '.join(row)
            # 优先识别晶盛光伏
            if "晶盛" in row_text and "光伏" in row_text:
                current_station = "晶盛光伏电站"
            else:
                station_match = re.search(r'机组[:：\s]*([^，。\n]+)', row_text) or re.search(r'(双发[AB]|双[AB])', row_text)
                if station_match:
                    current_station = station_match.group(1).strip()
            current_segment = [row]
        else:
            # 修复：保留阻塞/价差费用行
            if "阻塞费用" in row_str or "价差费用" in row_str:
                current_segment.append(row)
            else:
                current_segment.append(row)
    
    if current_segment:
        cleaned_station = clean_station_name(current_station)
        if cleaned_station == "未知场站":
            cleaned_station = extract_station_from_filename(file_name)
        station_segments.append((cleaned_station, current_segment))
    
    valid_segments = [(s, seg) for s, seg in station_segments if len(seg) >= 2 or s != "未知场站"]
    return valid_segments if valid_segments else [(extract_station_from_filename(file_name), merged_rows)]

def get_trade_name(trade_code, trade_text):
    """修复：强化阻塞/价差费用的名称匹配"""
    # 优先匹配编码
    if trade_code in TRADE_CODE_MAP:
        return TRADE_CODE_MAP[trade_code]
    # 强化关键词匹配
    trade_text_lower = trade_text.lower()
    for key, name in TRADE_KEYWORDS.items():
        if key in trade_text or key.lower() in trade_text_lower:
            return name
    # 兜底：直接匹配科目名称
    if "阻塞费用" in trade_text:
        return "中长期合约阻塞费用"
    if "价差费用" in trade_text:
        return "省间省内价差费用"
    return "未识别科目"

def parse_single_station_data(station_name, table_segment, company_name, clear_date):
    """核心修复：确保阻塞/价差费用不被跳过"""
    trade_records = []
    valid_rows = []
    
    for row in table_segment:
        if not row:
            continue
        row_clean = [remove_redundant_text(cell) for cell in row]
        row_str = ''.join(row_clean).replace(" ", "")
        is_empty = all(cell == '' for cell in row_clean)
        is_header = any(keyword in row_str for keyword in ["科目编码", "结算类型", "电量", "电价", "电费"])
        
        has_code = any(re.match(r'^\d{9,10}$', cell.replace(" ", "")) for cell in row_clean)
        has_trade_key = any(key in row_str for key in TRADE_KEYWORDS.keys())
        # 修复：单独判断阻塞/价差费用的有效数据
        has_block_spread = "阻塞费用" in row_str or "价差费用" in row_str
        has_valid_data = any(safe_convert_to_numeric(cell) is not None for cell in row_clean if cell not in ['', '-'])
        
        # 修复：阻塞/价差费用行直接判定为有效
        if (has_code or has_trade_key or has_valid_data or has_block_spread) and not is_empty and not is_header:
            valid_rows.append(row_clean)
    
    if len(valid_rows) < 1:  # 放宽行数要求
        return trade_records
    
    cols = {"code": -1, "name": -1, "qty": -1, "price": -1, "fee": -1}
    header_idx = -1
    for idx, row in enumerate(valid_rows[:3]):
        row_str = ''.join(row).replace(" ", "")
        if "结算类型" in row_str:
            header_idx = idx
            break
    header_idx = header_idx if header_idx != -1 else 0
    header_row = valid_rows[header_idx]
    
    for col_idx, cell in enumerate(header_row):
        cell_clean = remove_redundant_text(cell).lower()
        if "编码" in cell_clean:
            cols["code"] = col_idx
        elif "结算类型" in cell_clean:
            cols["name"] = col_idx
        elif "电量" in cell_clean and "价" not in cell_clean:
            cols["qty"] = col_idx
        elif "电价" in cell_clean or "单价" in cell_clean:
            cols["price"] = col_idx
        elif "电费" in cell_clean or "金额" in cell_clean:
            cols["fee"] = col_idx
    
    if any(v == -1 for v in cols.values()) and len(header_row) >= 5:
        cols = {"code": 0, "name": 1, "qty": 2, "price": 3, "fee": 4}
    
    data_start_idx = header_idx + 1
    for row_idx in range(data_start_idx, len(valid_rows)):
        row = valid_rows[row_idx]
        row_str = ''.join(row).replace(" ", "")
        is_subtotal = "小计" in row_str
        
        if "合计" in row_str and not is_subtotal:
            continue
        
        trade_code = row[cols["code"]].strip().replace(" ", "") if (cols["code"] < len(row)) else ""
        trade_text = row[cols["name"]].strip() if (cols["name"] < len(row)) else ""
        trade_name = get_trade_name(trade_code, trade_text)
        
        # 修复：即使是未识别科目，只要是阻塞/价差费用也保留
        if trade_name == "未识别科目":
            if "阻塞费用" in trade_text:
                trade_name = "中长期合约阻塞费用"
            elif "价差费用" in trade_text:
                trade_name = "省间省内价差费用"
            else:
                continue
        
        if is_subtotal:
            subtotal_qty = None
            subtotal_fee = None
            nums = [safe_convert_to_numeric(cell, "电量(兆瓦时)") for cell in row if isinstance(safe_convert_to_numeric(cell), (int, float))]
            fee_nums = [safe_convert_to_numeric(cell, "电费(元)") for cell in row if isinstance(safe_convert_to_numeric(cell), (int, float))]
            subtotal_qty = nums[0] if nums else None
            subtotal_fee = fee_nums[-1] if fee_nums else None
            trade_records.append({
                "公司名称": company_name,
                "场站名称": station_name,
                "清分日期": clear_date,
                "科目名称": "当日小计",
                "原始科目编码": trade_code,
                "原始科目文本": trade_text,
                "电量(兆瓦时)": subtotal_qty,
                "电价(元/兆瓦时)": None,
                "电费(元)": subtotal_fee
            })
            continue
        
        # 提取数据
        quantity = safe_convert_to_numeric(row[cols["qty"]], "电量(兆瓦时)") if (cols["qty"] < len(row)) else None
        price = safe_convert_to_numeric(row[cols["price"]], "电价(元/兆瓦时)") if (cols["price"] < len(row)) else None
        fee = safe_convert_to_numeric(row[cols["fee"]], "电费(元)") if (cols["fee"] < len(row)) else None
        
        # 修复：阻塞/价差费用特殊处理
        if "阻塞费用" in trade_name or "价差费用" in trade_name:
            quantity = None
            price = None
            # 即使fee为None也保留（避免跳过）
            trade_records.append({
                "公司名称": company_name,
                "场站名称": station_name,
                "清分日期": clear_date,
                "科目名称": trade_name,
                "原始科目编码": trade_code,
                "原始科目文本": trade_text,
                "电量(兆瓦时)": quantity,
                "电价(元/兆瓦时)": price,
                "电费(元)": fee
            })
        else:
            # 其他科目原有逻辑
            if quantity is None and fee is None:
                continue
            trade_records.append({
                "公司名称": company_name,
                "场站名称": station_name,
                "清分日期": clear_date,
                "科目名称": trade_name,
                "原始科目编码": trade_code,
                "原始科目文本": trade_text,
                "电量(兆瓦时)": quantity,
                "电价(元/兆瓦时)": price,
                "电费(元)": fee
            })
    
    return trade_records

# ---------------------- 主解析函数 ----------------------
def parse_pdf_final(file_obj, file_name):
    try:
        file_obj.seek(0)
        file_bytes = BytesIO(file_obj.read())
        file_bytes.seek(0)
        
        all_text = ""
        all_tables = []
        with pdfplumber.open(file_bytes) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"
                tables = page.extract_tables({
                    "vertical_strategy": "lines_strict",
                    "horizontal_strategy": "lines_strict",
                    "snap_tolerance": 1,
                    "join_tolerance": 1,
                    "edge_min_length": 3
                })
                all_tables.extend(tables)
        
        company_name = extract_company_info(all_text, file_name)
        clear_date = extract_clear_date(all_text, file_name)
        station_segments = split_double_station_tables(all_tables, all_text, file_name)
        
        all_records = []
        for station_name, table_segment in station_segments:
            final_station = clean_station_name(station_name)
            station_data = parse_single_station_data(final_station, table_segment, company_name, clear_date)
            all_records.extend(station_data)
        
        unique_records = []
        seen_keys = set()
        for rec in all_records:
            key = f"{rec['场站名称']}_{rec['科目名称']}_{rec['原始科目编码']}"
            if key not in seen_keys:
                seen_keys.add(key)
                unique_records.append(rec)
        
        if not unique_records:
            fallback_station = extract_station_from_filename(file_name)
            unique_records.append({
                "公司名称": company_name,
                "场站名称": fallback_station,
                "清分日期": clear_date,
                "科目名称": "无有效数据",
                "原始科目编码": "",
                "原始科目文本": "",
                "电量(兆瓦时)": None,
                "电价(元/兆瓦时)": None,
                "电费(元)": None
            })
        
        return unique_records
    
    except Exception as e:
        st.error(f"PDF解析错误: {str(e)}")
        fallback_station = extract_station_from_filename(file_name)
        return [{
            "公司名称": "未知发电公司",
            "场站名称": fallback_station,
            "清分日期": None,
            "科目名称": "解析失败",
            "原始科目编码": "",
            "原始科目文本": "",
            "电量(兆瓦时)": None,
            "电价(元/兆瓦时)": None,
            "电费(元)": None
        }]

# ---------------------- Streamlit应用 ----------------------
def main():
    st.set_page_config(page_title="通用日清分数据提取工具（晶盛+阻塞价差修复版）", layout="wide")
    
    st.title("📊 通用现货日清分结算单数据提取工具（精准完整版）")
    st.markdown("**核心修复：1.晶盛光伏站→晶盛光伏电站 2.阻塞/价差费用精准提取 3.负数电费正常解析**")
    st.divider()
    
    uploaded_files = st.file_uploader(
        "上传PDF文件（支持晶盛光伏/双发风电场，含阻塞/价差费用）",
        type=["pdf"],
        accept_multiple_files=True
    )
    
    if uploaded_files and st.button("🚀 开始批量提取", type="primary"):
        st.divider()
        st.subheader("⚙️ 处理进度")
        
        all_results = []
        progress_bar = st.progress(0)
        
        for idx, file in enumerate(uploaded_files):
            st.write(f"正在处理：{file.name}")
            file_results = parse_pdf_final(file, file.name)
            all_results.extend(file_results)
            progress_bar.progress((idx + 1) / len(uploaded_files))
            file.close()
        
        progress_bar.empty()
        
        df = pd.DataFrame(all_results).fillna("")
        display_cols = [
            "公司名称", "场站名称", "清分日期", "科目名称", 
            "原始科目编码", "原始科目文本", "电量(兆瓦时)", 
            "电价(元/兆瓦时)", "电费(元)"
        ]
        df_display = df[[col for col in display_cols if col in df.columns]]
        
        st.subheader("📈 批量提取结果（含阻塞/价差费用）")
        def highlight_rows(row):
            if row["科目名称"] == "当日小计":
                return ["background-color: #e6f3ff"] * len(row)
            elif row["场站名称"] == "双发A风电场":
                return ["background-color: #f0fff4"] * len(row)
            elif row["场站名称"] == "双发B风电场":
                return ["background-color: #fff8f0"] * len(row)
            elif row["场站名称"] == "晶盛光伏电站":
                return ["background-color: #f0f8ff"] * len(row)
            elif "阻塞费用" in row["科目名称"] or "价差费用" in row["科目名称"]:
                return ["background-color: #ffebee"] * len(row)  # 阻塞/价差费用高亮
            else:
                return [""] * len(row)
        styled_df = df_display.style.apply(highlight_rows, axis=1)
        st.dataframe(styled_df, use_container_width=True)
        
        total_stations = df["场站名称"].nunique()
        total_trades = len(df[(df["科目名称"] != "当日小计") & (df["科目名称"] != "无有效数据") & (df["科目名称"] != "解析失败")])
        block_spread_count = len(df[(df["科目名称"] == "中长期合约阻塞费用") | (df["科目名称"] == "省间省内价差费用")])
        subtotal_count = len(df[df["科目名称"] == "当日小计"])
        st.info(f"**统计：** 覆盖场站 {total_stations} 个 | 有效科目 {total_trades} 个 | 阻塞/价差费用 {block_spread_count} 个 | 小计行 {subtotal_count} 个")
        
        download_cols = [
            "公司名称", "场站名称", "清分日期", "科目名称", 
            "电量(兆瓦时)", "电价(元/兆瓦时)", "电费(元)"
        ]
        df_download = df[[col for col in download_cols if col in df.columns]]
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_download.to_excel(writer, index=False, sheet_name="多场站日清分数据")
            ws = writer.sheets["多场站日清分数据"]
            light_blue = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            light_red = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            for row in range(2, len(df_download) + 2):
                if df_download.iloc[row-2]["科目名称"] == "当日小计":
                    for col in range(1, len(df_download.columns) + 1):
                        ws.cell(row=row, column=col).fill = light_blue
                elif "阻塞费用" in df_download.iloc[row-2]["科目名称"] or "价差费用" in df_download.iloc[row-2]["科目名称"]:
                    for col in range(1, len(df_download.columns) + 1):
                        ws.cell(row=row, column=col).fill = light_red
        
        output.seek(0)
        st.download_button(
            label="📥 下载Excel（含阻塞/价差费用）",
            data=output,
            file_name=f"多场站日清分数据_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        
        st.success("✅ 提取完成！阻塞/价差费用已正常提取，晶盛光伏站名称已补全")
    
    else:
        st.info("👆 请上传晶盛光伏电站或双发A/B风电场的现货日清分结算单PDF（支持阻塞/价差费用提取）")

if __name__ == "__main__":
    os.environ["PYTHONIOENCODING"] = "utf-8"
    main()
