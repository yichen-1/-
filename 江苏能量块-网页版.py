import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from io import BytesIO

# -------------------------- 全局配置 --------------------------
ROUND_DECIMALS = 1  # 统一保留1位小数
FEB1_SHEET_NAME = "2.1"  # 功率文件中2月1日的sheet名（可根据需求修改）

# -------------------------- 核心计算函数 --------------------------
def get_position_data(uploaded_file):
    """读取持仓数据（统一24时段：0-23点）"""
    try:
        pos_df = pd.read_excel(uploaded_file, engine='openpyxl', header=0)
        if pos_df.shape[1] < 5:
            st.error("持仓文件列数不足，需至少5列（E列存储持仓数据）")
            return None
        
        positions = []
        for val in pos_df.iloc[:, 4]:  # 读取E列（索引4）
            positions.append(round(float(val), ROUND_DECIMALS) if pd.notna(val) else 0.0)
        positions = positions[:24] + [0.0] * (24 - len(positions))
        return positions
    except Exception as e:
        st.error(f"持仓文件读取失败：{str(e)}")
        return None

def get_valid_power_data(uploaded_file, sheet_name):
    """读取功率数据，统一映射到0-23时段（4-20点有效，其余补0）"""
    try:
        # 读取功率文件（支持.xls格式）
        df = pd.read_excel(
            uploaded_file,
            sheet_name=sheet_name,
            engine='xlrd',
            usecols=[1],  # 只读取B列功率值
            skiprows=1,
            header=None
        )
        df.columns = ["原始功率(kW)"]
        df["原始功率(kW)"] = pd.to_numeric(df["原始功率(kW)"], errors='coerce').fillna(0)

        # 补全96个15分钟数据点
        if len(df) < 96:
            pad = pd.DataFrame({"原始功率(kW)": [0.0] * (96 - len(df))})
            df = pd.concat([df, pad], ignore_index=True)
        df = df.head(96)

        # 映射到统一24时段
        period_power = [0.0] * 24
        valid_start_period = 4  # 有效数据从4点开始
        valid_period_count = 17  # 4-20点共17个时段

        for i in range(valid_period_count):
            if valid_start_period + i < 24:
                power_avg = df["原始功率(kW)"].iloc[i*4 : (i+1)*4].mean()
                period_power[valid_start_period + i] = round(power_avg, ROUND_DECIMALS)
        return period_power
    except Exception as e:
        st.warning(f"功率sheet【{sheet_name}】处理失败，使用全0数据：{str(e)}")
        return [0.0] * 24

def calc_unified_balance(daily_power, positions, feb1_power):
    """计算差额（统一保留1位小数）"""
    daily_01 = [round(p * 0.1, ROUND_DECIMALS) for p in daily_power]
    daily_balance = [round(d01 - pos, ROUND_DECIMALS) for d01, pos in zip(daily_01, positions)]

    feb1_01 = [round(p * 0.1, ROUND_DECIMALS) for p in feb1_power]
    feb1_balance = [round(f01 - pos, ROUND_DECIMALS) for f01, pos in zip(feb1_01, positions)]

    final_balance = [round(d_bal - f_bal, ROUND_DECIMALS) for d_bal, f_bal in zip(daily_balance, feb1_balance)]
    return daily_power, daily_01, final_balance

def generate_excel_with_highlight(df):
    """生成带负差额标黄的Excel文件（BytesIO格式）"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="统一24时段汇总", index=False)

    # 打开工作簿，为负差额标黄
    wb = writer.book
    ws = wb["统一24时段汇总"]
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 遍历所有列，找到差额列并标黄
    for col_idx in range(1, ws.max_column + 1):
        col_name = ws.cell(row=1, column=col_idx).value
        if col_name and "差额" in str(col_name):
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if isinstance(val, (int, float)) and val < 0:
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

    wb.save(output)
    output.seek(0)
    return output

# -------------------------- Streamlit 页面 --------------------------
st.set_page_config(page_title="功率持仓计算工具", layout="wide")
st.title("⚡ 功率-持仓统一时段计算工具")
st.markdown("""
**使用说明**：
1. 上传功率文件（格式：.xls，包含多个日期sheet）
2. 上传持仓文件（格式：.xlsx，E列存储24时段持仓数据）
3. 自动计算并展示结果，支持下载带标黄的Excel文件
""")

# 文件上传区域
col1, col2 = st.columns(2)
with col1:
    power_file = st.file_uploader("上传功率文件（.xls）", type=["xls"])
with col2:
    position_file = st.file_uploader("上传持仓文件（.xlsx）", type=["xlsx"])

# 计算按钮与结果展示
if st.button("🚀 开始计算") and power_file and position_file:
    with st.spinner("正在处理数据..."):
        # 1. 读取持仓数据
        positions = get_position_data(position_file)
        if positions is None:
            st.stop()

        # 2. 读取功率文件所有sheet
        try:
            power_xls = pd.ExcelFile(power_file, engine='xlrd')
            all_dates = power_xls.sheet_names
            st.success(f"✅ 检测到功率文件共 {len(all_dates)} 个日期：{all_dates}")
        except Exception as e:
            st.error(f"功率文件读取失败：{str(e)}")
            st.stop()

        # 3. 计算2月1日基准数据
        feb1_power = get_valid_power_data(power_file, FEB1_SHEET_NAME)

        # 4. 初始化结果表
        summary_data = {
            "统一时段（点）": list(range(24)),
            "持仓值(kWh)": positions
        }

        # 5. 遍历所有日期计算
        for date in all_dates:
            daily_power = get_valid_power_data(power_file, date)
            daily_power, daily_01, final_balance = calc_unified_balance(daily_power, positions, feb1_power)
            summary_data[f"{date}_发电量(kWh)"] = daily_power
            summary_data[f"{date}_0.1倍发电量(kWh)"] = daily_01
            summary_data[f"{date}_差额(kWh)"] = final_balance

        # 6. 生成结果DataFrame
        result_df = pd.DataFrame(summary_data)

        # 7. 展示结果（Streamlit表格）
        st.subheader("📊 计算结果预览")
        st.dataframe(result_df, use_container_width=True)

        # 8. 生成带标黄的Excel并提供下载
        excel_file = generate_excel_with_highlight(result_df)
        st.download_button(
            label="📥 下载结果Excel（带负差额标黄）",
            data=excel_file,
            file_name="功率持仓计算结果.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.success("🎉 计算完成！所有结果已保留1位小数，负差额自动标黄")

elif st.button("🚀 开始计算") and (not power_file or not position_file):
    st.warning("⚠️ 请先上传功率文件和持仓文件！")
